"""
app.py  ─  DWSIM Cost Estimation Web Application
Powered by Streamlit | Turton et al. (2012) cost correlations
Run:  streamlit run app.py
"""
from __future__ import annotations

import io
import math
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# Ensure sibling modules are importable when run from any cwd
sys.path.insert(0, str(Path(__file__).parent))

from excel_parser import DWSIMData
from cost_engine import CostEngine, CostSettings, CostReport
from equipment_models import UTILITY_COSTS, CEPCI_2024, CEPCI_2001

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="DWSIM Cost Estimator",
    page_icon="⚗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Custom CSS – premium dark-gradient theme
# ---------------------------------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@400;600;700&display=swap');

/* ── Global ── */
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp {
    background: linear-gradient(135deg, #0f0c29 0%, #1a1a3e 40%, #24243e 100%);
    color: #e0e0ff;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1a3e 0%, #12122a 100%);
    border-right: 1px solid rgba(130,100,255,0.2);
}
[data-testid="stSidebar"] .block-container { padding-top: 1rem; }

/* ── Hero banner ── */
.hero-banner {
    background: linear-gradient(120deg, #4f2af5 0%, #8b5cf6 50%, #06b6d4 100%);
    border-radius: 18px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 8px 40px rgba(79,42,245,0.4);
    position: relative;
    overflow: hidden;
}
.hero-banner::before {
    content:'';
    position:absolute; top:-40px; right:-40px;
    width:200px; height:200px;
    background: rgba(255,255,255,0.05);
    border-radius:50%;
}
.hero-banner h1 {
    font-family: 'Outfit', sans-serif;
    font-size: 2.2rem; font-weight: 700;
    margin: 0 0 0.3rem 0; color: #fff;
    text-shadow: 0 2px 10px rgba(0,0,0,0.3);
}
.hero-banner p { margin:0; font-size:1.05rem; color:rgba(255,255,255,0.85); }

/* ── Metric cards ── */
.metric-grid { display: flex; gap: 1rem; flex-wrap: wrap; margin-bottom: 1.5rem; }
.metric-card {
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(139,92,246,0.3);
    border-radius: 14px;
    padding: 1.1rem 1.4rem;
    flex: 1; min-width: 160px;
    backdrop-filter: blur(10px);
    transition: border-color 0.3s, transform 0.2s;
}
.metric-card:hover { border-color: rgba(139,92,246,0.7); transform: translateY(-2px); }
.metric-card .label { font-size: 0.76rem; color: #a0a0cc; text-transform:uppercase; letter-spacing:0.06em; }
.metric-card .value { font-size: 1.55rem; font-weight:700; color: #c4b5fd; margin:0.2rem 0 0 0; }
.metric-card .sub   { font-size: 0.78rem; color:#7c7ca0; }

/* ── Section headers ── */
.section-header {
    display: flex; align-items: center; gap: 0.6rem;
    font-family: 'Outfit', sans-serif;
    font-size: 1.25rem; font-weight: 600; color: #c4b5fd;
    margin: 1.8rem 0 0.8rem 0;
    padding-bottom: 0.4rem;
    border-bottom: 2px solid rgba(139,92,246,0.3);
}

/* ── Tables ── */
.dataframe { background: rgba(255,255,255,0.03) !important; border-radius:10px; }
thead tr th {
    background: linear-gradient(90deg,#4f2af5,#8b5cf6) !important;
    color: #fff !important; font-weight:600 !important;
}
tbody tr:hover td { background: rgba(139,92,246,0.08) !important; }

/* ── Warning pills ── */
.warn-pill {
    background: rgba(251,191,36,0.15);
    border: 1px solid rgba(251,191,36,0.4);
    border-radius: 8px; padding: 0.35rem 0.7rem;
    font-size:0.82rem; color:#fbbf24; margin:0.25rem 0;
    display: inline-block;
}

/* ── Upload zone ── */
[data-testid="stFileUploadDropzone"] {
    border: 2px dashed rgba(139,92,246,0.5) !important;
    border-radius: 14px !important;
    background: rgba(139,92,246,0.05) !important;
}

/* ── Tabs ── */
button[data-baseweb="tab"] {
    font-family:'Outfit',sans-serif; font-weight:600;
    color: #9090c0 !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #c4b5fd !important;
    border-bottom: 3px solid #8b5cf6 !important;
}

/* ── Buttons ── */
.stDownloadButton > button, .stButton > button {
    background: linear-gradient(135deg, #4f2af5, #8b5cf6) !important;
    color: #fff !important; border: none !important;
    border-radius: 10px !important; font-weight: 600 !important;
    padding: 0.5rem 1.5rem !important;
    transition: opacity 0.2s, transform 0.2s !important;
}
.stDownloadButton > button:hover, .stButton > button:hover {
    opacity: 0.88; transform: translateY(-1px);
}

/* ── Sliders / inputs ── */
[data-testid="stSlider"] .css-1vy1n3z { color:#c4b5fd; }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Helper: format currency
# ---------------------------------------------------------------------------
def fmt_usd(val: float, suffix: str = "") -> str:
    symbol = "$"
    if "settings" in st.session_state and st.session_state.settings:
        sym = st.session_state.settings.currency_symbol
        if sym and sym != "USD":
            symbol = sym + " "
    if val >= 1e9:
        return f"{symbol}{val/1e9:.3f}B{suffix}"
    if val >= 1e6:
        return f"{symbol}{val/1e6:.3f}M{suffix}"
    if val >= 1e3:
        return f"{symbol}{val/1e3:.1f}K{suffix}"
    return f"{symbol}{val:.0f}{suffix}"



def fmt_pct(val: float) -> str:
    return f"{val:.1f}%"


# ---------------------------------------------------------------------------
# Raw Material Heuristics & Database Loader
# ---------------------------------------------------------------------------
def detect_feed_streams(data: DWSIMData) -> list[str]:
    if data.streams.empty:
        return []
    material_streams = data.streams[data.streams["ObjectType"] == "Material Stream"]["Tag"].tolist()
    if data.connections.empty:
        return material_streams
        
    produced_streams = set()
    cols = [c.lower() for c in data.connections.columns]
    
    if "outputstreams" in cols:
        for _, row in data.connections.iterrows():
            outs = row.get("OutputStreams") or row.get("outputstreams")
            if outs:
                for part in str(outs).split("|"):
                    produced_streams.add(part.strip())
    elif "from" in cols and "to" in cols:
        for _, row in data.connections.iterrows():
            src = str(row.get("From") or row.get("from", "")).strip()
            dest = str(row.get("To") or row.get("to", "")).strip()
            if dest in material_streams and src not in material_streams:
                produced_streams.add(dest)
                
    feeds = [s for s in material_streams if s not in produced_streams]
    if not feeds:
        return material_streams
    return feeds


def load_raw_material_db() -> pd.DataFrame:
    path = r"C:\Users\Dr. Shephali\Downloads\dwsim_raw_material_costs.csv"
    if os.path.exists(path):
        try:
            return pd.read_csv(path)
        except Exception:
            pass
    return pd.DataFrame()


def lookup_price(compound_name: str, pricing_dict: dict[str, float]) -> float:
    c_low = compound_name.lower().strip()
    if c_low in pricing_dict:
        return pricing_dict[c_low]
    # Try matching by words
    c_words = set(c_low.split())
    for k, v in pricing_dict.items():
        k_words = set(k.split())
        if c_words == k_words:
            return v
    return 0.0


def calculate_raw_materials(data: DWSIMData, db_df: pd.DataFrame, selected_feeds: list[str], op_hours: float) -> tuple[pd.DataFrame, float]:
    if data.streams.empty or data.compositions.empty or db_df.empty:
        return pd.DataFrame(), 0.0
        
    pricing = {}
    for _, row in db_df.iterrows():
        name = str(row.get("name", "")).strip().lower()
        price = row.get("price_usd_per_kg")
        try:
            if name and not pd.isna(price):
                pricing[name] = float(price)
        except ValueError:
            pass
            
    flow_map = {}
    for _, row in data.streams.iterrows():
        tag = str(row.get("Tag", "")).strip()
        flow = row.get("MassFlow_kg_s")
        try:
            if tag and flow is not None:
                flow_map[tag] = float(flow)
        except ValueError:
            pass
            
    comp_df = data.compositions
    tag_col = "StreamTag" if "StreamTag" in comp_df.columns else "Tag"
    phase_col = "Phase" if "Phase" in comp_df.columns else None
    
    if phase_col and phase_col in comp_df.columns:
        has_mixture = comp_df[phase_col].astype(str).str.lower().str.strip().eq("mixture").any()
        if has_mixture:
            comp_df = comp_df[comp_df[phase_col].astype(str).str.lower().str.strip().eq("mixture")]
        else:
            has_overall = comp_df[phase_col].astype(str).str.lower().str.strip().eq("overall").any()
            if has_overall:
                comp_df = comp_df[comp_df[phase_col].astype(str).str.lower().str.strip().eq("overall")]
                
    rows = []
    total_cost = 0.0
    
    for _, row in comp_df.iterrows():
        stream = str(row.get(tag_col, "")).strip()
        if stream not in selected_feeds:
            continue
        compound = str(row.get("Compound", "")).strip()
        
        try:
            mass_frac = float(row.get("MassFraction", 0.0))
        except ValueError:
            mass_frac = 0.0
            
        stream_flow = flow_map.get(stream, 0.0)
        comp_flow_kg_s = stream_flow * mass_frac
        comp_flow_kg_yr = comp_flow_kg_s * 3600 * op_hours
        
        price = lookup_price(compound, pricing)
        cost_yr = comp_flow_kg_yr * price
        total_cost += cost_yr
        
        rows.append({
            "Feed Stream": stream,
            "Compound": compound,
            "Mass Flow (kg/s)": comp_flow_kg_s,
            "Mass Flow (kg/yr)": comp_flow_kg_yr,
            "Price ($/kg)": price,
            "Annual Cost ($/yr)": cost_yr
        })
        
    return pd.DataFrame(rows), total_cost


# ---------------------------------------------------------------------------
# Session state defaults
# ---------------------------------------------------------------------------
if "data" not in st.session_state:
    st.session_state.data = None
if "report" not in st.session_state:
    st.session_state.report = None
if "utility_database" not in st.session_state:
    from cost_engine import DEFAULT_UTILITY_DATABASE
    st.session_state.utility_database = {k: dict(v) for k, v in DEFAULT_UTILITY_DATABASE.items()}
if "settings" not in st.session_state:
    st.session_state.settings = CostSettings(utility_database=st.session_state.utility_database)


# ---------------------------------------------------------------------------
# Sidebar – Settings
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.markdown("---")

    st.markdown("### 📅 Cost Index")
    cepci = st.number_input("CEPCI (current year)", value=780.0, min_value=200.0,
                             max_value=2000.0, step=5.0,
                             help="Chemical Engineering Plant Cost Index. 2024 ≈ 780, 2001 = 397 (Turton base)")

    st.markdown("### 🏗️ CAPEX")
    capex_mode = st.selectbox(
        "CAPEX Mode",
        options=["Rigorous (System Expansion)", "Strict (Legacy Turton)"],
        index=0,
        help="Rigorous sizes distillation auxiliaries and uses grassroots roll-up. Strict costs columns only."
    )

    mat_factor = st.selectbox(
        "Material factor (F_M)",
        options=[("Carbon Steel (1.0)", 1.0),
                 ("Stainless Steel 304 (1.7)", 1.7),
                 ("Stainless Steel 316 (2.1)", 2.1),
                 ("Hastelloy C (2.9)", 2.9),
                 ("Titanium (4.0)", 4.0)],
        format_func=lambda x: x[0],
        index=0,
    )[1]

    location_factor = st.number_input(
        "Location Factor (L_F)",
        value=1.00,
        min_value=0.1,
        max_value=10.0,
        step=0.05,
        help="Regional/location cost factor (e.g. USGC = 1.00)"
    )
    
    currency_symbol = st.text_input(
        "Currency Symbol",
        value="USD",
        help="Symbol to display costs in"
    )
    
    currency_conversion = st.number_input(
        "Currency Conversion Factor",
        value=1.00,
        min_value=0.001,
        max_value=100000.0,
        step=0.1,
        help="USD to local currency rate (e.g. 1 USD = X Local)"
    )

    contingency_pct = st.slider("Contingency (%)", 5.0, 30.0, 15.0, 1.0)
    engineering_pct = st.slider("Engineering fee (%)", 5.0, 20.0, 10.0, 1.0)
    startup_pct     = st.slider("Startup cost (%)", 1.0, 15.0, 5.0, 0.5)
    working_cap_pct = st.slider("Working capital (%)", 5.0, 25.0, 10.0, 1.0)

    st.markdown("### ⚡ OPEX")
    opex_method = st.selectbox(
        "OPEX Estimation Method", 
        ["Turton (COMd) - Rigorous", "Simple Addition - Basic"], 
        index=0, 
        help="Turton uses rigorous multipliers: COMd = 0.18*FCI + 2.73*Labor + 1.23*(Utilities + RawMaterials)"
    )
    if "raw_material_cost" not in st.session_state:
        st.session_state.raw_material_cost = 0.0
    raw_material_cost = st.number_input(
        "Annual Raw Material Cost (USD/yr)",
        value=float(st.session_state.raw_material_cost),
        min_value=0.0,
        step=100000.0,
    )
    st.session_state.raw_material_cost = raw_material_cost
    op_hours = st.number_input("Operating hours/year", value=8000, min_value=1000,
                                max_value=8760, step=100)
    elec_price = st.number_input("Electricity price (USD/kWh)", value=0.06,
                                  min_value=0.01, max_value=0.50, step=0.01)
    steam_util = st.selectbox("Reboiler/heater utility",
                               options=["LP Steam (3 bar)", "MP Steam (10 bar)", "HP Steam (42 bar)"],
                               index=1)
    cool_util  = st.selectbox("Condenser/cooler utility",
                               options=["Cooling Water", "Chilled Water", "Refrigerant (-25°C)"],
                               index=0)
    labour_cost = st.number_input("Annual labour cost (USD/yr)",
                                   value=0, min_value=0, step=50_000, help="Leave as 0 to use automated heuristics.")
    
    maintenance_pct = st.number_input("Maintenance cost (% of FCI/yr)",
                                       value=float(st.session_state.settings.maintenance_pct),
                                       min_value=0.0, max_value=20.0, step=0.5,
                                       help="Turton standard is 6.0%")
    
    insurance_pct = st.number_input("Insurance & Tax (% of FCI/yr)",
                                     value=float(st.session_state.settings.insurance_pct),
                                     min_value=0.0, max_value=10.0, step=0.5,
                                     help="Turton standard is 3.0%")

    st.markdown("### 📊 Profitability")
    plant_life  = st.number_input("Plant life (years)", value=10, min_value=1, max_value=40, step=1)
    disc_rate   = st.slider("Discount rate (%)", 1.0, 20.0, 10.0, 0.5)

    with st.expander("🛠️ Equipment Construction Materials", expanded=False):
        pump_material = st.selectbox(
            "Pump Material",
            options=[("Carbon Steel", "CS_carbon_steel"), ("Cast Iron", "Fe_cast_iron"), ("Stainless Steel", "SS_stainless_steel"), ("Nickel Alloy", "Ni_nickel_alloy"), ("Titanium", "Ti_titanium")],
            format_func=lambda x: x[0],
            index=0
        )[1]

        hx_material = st.selectbox(
            "Heat Exchanger Material",
            options=[("Carbon Steel / Carbon Steel", "CS/CS"), ("Carbon Steel / Stainless Steel", "CS/SS"), ("Stainless Steel / Stainless Steel", "SS/SS"), ("Carbon Steel / Nickel Alloy", "CS/Ni"), ("Nickel Alloy / Nickel Alloy", "Ni/Ni"), ("Carbon Steel / Titanium", "CS/Ti"), ("Titanium / Titanium", "Ti/Ti")],
            format_func=lambda x: x[0],
            index=0
        )[1]

        vessel_material = st.selectbox(
            "Vessel/Reactor Material",
            options=[("Carbon Steel", "CS_carbon_steel"), ("Stainless Steel", "SS_stainless_steel"), ("Nickel Alloy", "Ni_nickel_alloy"), ("Titanium", "Ti_titanium")],
            format_func=lambda x: x[0],
            index=0
        )[1]

        compressor_material = st.selectbox(
            "Compressor Material",
            options=[("Carbon Steel", "CS_carbon_steel"), ("Stainless Steel", "SS_stainless_steel"), ("Nickel Alloy", "Ni_nickel_alloy")],
            format_func=lambda x: x[0],
            index=0
        )[1]

        tray_material = st.selectbox(
            "Column Tray Material",
            options=[("Carbon Steel", "CS_carbon_steel"), ("Stainless Steel", "SS_stainless_steel"), ("Nickel Alloy", "Ni_nickel_alloy")],
            format_func=lambda x: x[0],
            index=0
        )[1]

        tray_type = st.selectbox(
            "Column Tray Type",
            options=[("Sieve Tray", "sieve"), ("Valve Tray", "valve")],
            format_func=lambda x: x[0],
            index=0
        )[1]

    # Configurable Utility Database overrides
    with st.expander("🔌 Configurable Utility Settings", expanded=False):
        if "utility_database" not in st.session_state:
            from cost_engine import DEFAULT_UTILITY_DATABASE
            st.session_state.utility_database = {k: dict(v) for k, v in DEFAULT_UTILITY_DATABASE.items()}
            
        for ut_name, ut_info in list(st.session_state.utility_database.items()):
            st.markdown(f"**{ut_name}**")
            cols = st.columns(3)
            if ut_info["type"] == "cooling":
                ut_info["Tin_C"] = cols[0].number_input(f"Tin (°C)", value=float(ut_info["Tin_C"]), key=f"ut_{ut_name}_Tin", step=1.0)
                ut_info["Tout_C"] = cols[1].number_input(f"Tout (°C)", value=float(ut_info["Tout_C"]), key=f"ut_{ut_name}_Tout", step=1.0)
            else:
                ut_info["Tsat_C"] = cols[0].number_input(f"Tsat (°C)", value=float(ut_info["Tsat_C"]), key=f"ut_{ut_name}_Tsat", step=1.0)
            
            ut_info["U_W_m2K"] = cols[2].number_input(f"U (W/m²K)", value=float(ut_info["U_W_m2K"]), key=f"ut_{ut_name}_U", step=10.0)
            ut_info["price_USD_GJ"] = st.number_input(f"Price ($/GJ)", value=float(ut_info["price_USD_GJ"]), key=f"ut_{ut_name}_price", step=0.1)

    # Reflux drum sizing parameters overrides
    with st.expander("⚗️ Reflux Drum Sizing Settings", expanded=False):
        reflux_drum_residence_time_s = st.number_input("Residence Time (s)", value=300.0, step=10.0, help="Turton standard is 300 s")
        reflux_drum_surge_factor = st.number_input("Surge Factor", value=2.0, step=0.1, help="Turton standard is 2.0")

    # Update settings
    s = CostSettings(
        cepci=cepci,
        material_factor=mat_factor,
        operating_hours=float(op_hours),
        elec_price_per_kwh=elec_price,
        steam_utility=steam_util,
        cooling_utility=cool_util,
        contingency_pct=contingency_pct,
        engineering_pct=engineering_pct,
        startup_pct=startup_pct,
        working_capital_pct=working_cap_pct,
        opex_method=opex_method.split(" -")[0],
        raw_material_cost=float(raw_material_cost),
        labour_cost_override=float(labour_cost),
        plant_life_years=float(plant_life),
        discount_rate_pct=disc_rate,
        maintenance_pct=maintenance_pct,
        insurance_pct=insurance_pct,
        location_factor=float(location_factor),
        currency_conversion=float(currency_conversion),
        currency_symbol=currency_symbol,
        capex_mode=capex_mode,
        pump_material=pump_material,
        hx_material=hx_material,
        vessel_material=vessel_material,
        compressor_material=compressor_material,
        tray_material=tray_material,
        tray_type=tray_type,
        reflux_drum_residence_time_s=float(reflux_drum_residence_time_s),
        reflux_drum_surge_factor=float(reflux_drum_surge_factor),
        utility_database=st.session_state.utility_database,
    )

    if st.session_state.data is not None:
        if st.button("🔄 Recalculate", use_container_width=True):
            engine = CostEngine(st.session_state.data, s)
            st.session_state.report = engine.estimate()
            st.session_state.settings = s
            st.rerun()


# ---------------------------------------------------------------------------
# Hero banner
# ---------------------------------------------------------------------------
st.markdown("""
<div class="hero-banner">
  <h1>⚗️ DWSIM Cost Estimator</h1>
  <p>Free CAPEX &amp; OPEX estimation powered by <strong>Turton et al. (2012)</strong> cost correlations
  &nbsp;·&nbsp; Upload your DWSIM introspection Excel and get instant cost breakdowns.</p>
</div>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Upload section
# ---------------------------------------------------------------------------
st.markdown('<div class="section-header">📂 Upload DWSIM File</div>',
            unsafe_allow_html=True)

sample_path = Path(__file__).parent.parent / "DWSIM_Costing_Introspection.xlsx"

# Two-tab upload interface: DWSIM file OR introspection Excel
up_tab1, up_tab2 = st.tabs(["🧪 Upload DWSIM Flowsheet (.dwxmz)", "📊 Upload Introspection Excel (.xlsx)"])

uploaded = None   # will be set to a BytesIO of the Excel by either path

# --- Path A: Direct .dwxmz upload ---
with up_tab1:
    st.markdown(
        "Upload your DWSIM flowsheet directly. "
        "The app will extract variables automatically using `costing_engine_variables.py`."
    )
    dwxmz_file = st.file_uploader(
        "DWSIM flowsheet",
        type=["dwxmz"],
        key="dwxmz_uploader",
        label_visibility="collapsed",
    )
    _nocalc = st.checkbox("Skip re-calculation (use pre-calculated values)", value=True,
                          help="Recommended for large flowsheets — avoids slow solver re-run.")
    _depth  = st.selectbox("Extraction depth", [2, 3, 4], index=0,
                           help="Depth 2 is fast and sufficient for costing; depth 4 is thorough but slow.")

    if dwxmz_file is not None:
        extractor_py = Path(__file__).parent.parent / "costing_engine_variables.py"
        if not extractor_py.exists():
            st.error(f"❌ Extractor not found at `{extractor_py}`. Check your project layout.")
        else:
            with st.spinner("⏳ Extracting variables from DWSIM flowsheet… this may take 30–90 s"):
                try:
                    # Save uploaded .dwxmz to a temp file
                    with tempfile.TemporaryDirectory() as tmpdir:
                        dwxmz_path = os.path.join(tmpdir, dwxmz_file.name)
                        with open(dwxmz_path, "wb") as fout:
                            fout.write(dwxmz_file.read())

                        xlsx_path = os.path.splitext(dwxmz_path)[0] + "_costing_introspection.xlsx"
                        cmd = [
                            sys.executable,
                            str(extractor_py),
                            dwxmz_path,
                            "--output", xlsx_path,
                            "--depth", str(_depth),
                        ]
                        if _nocalc:
                            cmd.append("--no-calc")

                        result = subprocess.run(
                            cmd,
                            capture_output=True, text=True, timeout=300,
                        )

                        if result.returncode != 0:
                            st.error("❌ Extraction failed. See details below.")
                            with st.expander("Extractor stderr", expanded=True):
                                st.code(result.stderr[-3000:] if result.stderr else "(no output)")
                        else:
                            if os.path.exists(xlsx_path):
                                with open(xlsx_path, "rb") as f:
                                    buf_dwxmz = io.BytesIO(f.read())
                                buf_dwxmz.name = os.path.basename(xlsx_path)
                                uploaded = buf_dwxmz
                                # Show extractor log in an expander
                                if result.stdout:
                                    with st.expander("📋 Extractor log", expanded=False):
                                        st.code(result.stdout[-3000:])
                            else:
                                st.error("❌ Extractor ran but produced no Excel file.")
                                st.code(result.stdout[-2000:])
                except subprocess.TimeoutExpired:
                    st.error("⏱️ Extraction timed out (>5 min). Try enabling 'Skip re-calculation'.")
                except Exception as exc:
                    st.error(f"❌ Unexpected error: {exc}")
                    st.exception(exc)

# --- Path B: Manual Excel upload ---
with up_tab2:
    st.markdown(
        "Upload the `*_costing_introspection.xlsx` file generated by `costing_engine_variables.py`."
    )
    col_up1, col_up2 = st.columns([3, 1])
    with col_up1:
        xlsx_uploaded = st.file_uploader(
            "Introspection Excel",
            type=["xlsx"],
            key="xlsx_uploader",
            label_visibility="collapsed",
        )
    with col_up2:
        if sample_path.exists():
            if st.button("📋 Load Sample File", use_container_width=True,
                         help="Load the bundled DWSIM_Costing_Introspection.xlsx"):
                with open(sample_path, "rb") as f:
                    buf_sample = io.BytesIO(f.read())
                buf_sample.name = sample_path.name  # type: ignore[attr-defined]
                xlsx_uploaded = buf_sample
    if xlsx_uploaded is not None:
        uploaded = xlsx_uploaded


if uploaded is not None:
    with st.spinner("Parsing flowsheet data…"):
        try:
            data = DWSIMData(uploaded if isinstance(uploaded, Path) else uploaded)
            
            # AUTO CALC RAW MATERIALS FROM DATABASE
            db_df = load_raw_material_db()
            if not db_df.empty:
                feeds = detect_feed_streams(data)
                st.session_state.selected_feeds = feeds
                _, calc_total = calculate_raw_materials(data, db_df, feeds, s.operating_hours)
                st.session_state.raw_material_cost = calc_total
                s.raw_material_cost = calc_total
                
            engine = CostEngine(data, s)
            report = engine.estimate()
            st.session_state.data = data
            st.session_state.report = report
            st.session_state.settings = s
            st.success(f"✅ Loaded: **{data.flowsheet_name}** · "
                       f"{len(data.equipment_inventory())} equipment items found")
        except Exception as exc:
            st.error(f"❌ Could not parse file: {exc}")
            st.exception(exc)


# ---------------------------------------------------------------------------
# Main content – only shown after file loaded
# ---------------------------------------------------------------------------
if st.session_state.data is None:
    st.info("⬆️ Upload a DWSIM flowsheet (.dwxmz) or introspection Excel (.xlsx) to begin, or click **Load Sample File**.")
    st.markdown("---")
    st.markdown("""
    ### 🚀 Two ways to start
    **Option A — Direct upload (recommended)**
    1. Go to the **Upload DWSIM Flowsheet** tab above
    2. Upload your `.dwxmz` file — variables are extracted automatically

    **Option B — Manual extraction**
    1. Run `costing_engine_variables.py` from the command line:
       ```
       python costing_engine_variables.py "your_flowsheet.dwxmz" --no-calc --depth 2
       ```
    2. Upload the resulting `*_costing_introspection.xlsx` in the Excel tab

    ### 📐 Costing Method
    | Layer | Method |
    |---|---|
    | Purchase cost | Turton et al. (2012) log-quadratic correlations |
    | Installation | Bare-module factor (F_BM) per equipment class |
    | ISBL | Bare-module + Contingency + Engineering |
    | OSBL | 10% of ISBL |
    | OPEX | Utility rate × annual duty + Labour + Maintenance |
    """)
    st.stop()


data: DWSIMData = st.session_state.data
report: CostReport = st.session_state.report

# ---------------------------------------------------------------------------
# KPI cards row
# ---------------------------------------------------------------------------
st.markdown("---")
c1, c2, c3, c4, c5 = st.columns(5)

def kpi(col, label, value, sub=""):
    col.markdown(f"""
    <div class="metric-card">
      <div class="label">{label}</div>
      <div class="value">{value}</div>
      <div class="sub">{sub}</div>
    </div>
    """, unsafe_allow_html=True)

kpi(c1, "Total Project Cost", fmt_usd(report.total_project_cost), "all-in")
kpi(c2, "Fixed Capital (ISBL+OSBL)", fmt_usd(report.total_fixed_capital), "")
kpi(c3, "Total OPEX", fmt_usd(report.total_opex, "/yr"), "annual")
kpi(c4, "Equipment Items", str(len(report.equipment_costs)), "priced")
kpi(c5, "Total Annual Cost", fmt_usd(report.total_annual_cost, "/yr"), "incl. ann. CAPEX")

# Unified Warnings & Validation Panel
if report.warnings:
    st.markdown('<div class="section-header">⚠️ Process & Cost Validation Panel</div>', unsafe_allow_html=True)
    errors = [w for w in report.warnings if w.startswith("❌")]
    warnings = [w for w in report.warnings if not w.startswith("❌")]
    
    if errors:
        for err in errors:
            st.error(err)
            
    if warnings:
        with st.expander(f"⚠️ {len(warnings)} Sizing & Validation Alerts", expanded=True):
            for w in warnings:
                st.markdown(f"- {w}")

# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🏭 Equipment Inventory",
    "💰 CAPEX Breakdown",
    "⚡ OPEX & Utilities",
    "📊 Cost Dashboard",
    "📄 Export Report",
])


# ============================================================
# Tab 1 – Equipment Inventory
# ============================================================
with tab1:
    st.markdown('<div class="section-header">🏭 Equipment Inventory</div>', unsafe_allow_html=True)

    inv = data.equipment_inventory()
    if inv:
        inv_df = pd.DataFrame(inv)
        display_cols = [c for c in ["tag", "type", "status"] if c in inv_df.columns]
        # Add key sizing params
        def size_summary(row):
            n_stages_raw = row.get('n_stages')
            try:
                n_stages_int = int(float(n_stages_raw)) if n_stages_raw is not None and str(n_stages_raw) not in ('', 'nan', 'None') else 0
            except (ValueError, TypeError):
                n_stages_int = 0
            if row.get("diameter_m") and str(row.get("diameter_m")) not in ('', 'nan', 'None'):
                d = float(row.get('diameter_m', 0) or 0)
                h = float(row.get('height_m', 0) or 0)
                return f"D={d:.2f}m, H={h:.1f}m, N={n_stages_int} stages"
            if row.get("area_m2") and str(row.get("area_m2")) not in ('', 'nan', 'None'):
                return f"A={float(row.get('area_m2', 0) or 0):.1f} m²"
            if row.get("power_kw") and str(row.get("power_kw")) not in ('', 'nan', 'None'):
                return f"P={float(row.get('power_kw', 0) or 0):.1f} kW"
            if row.get("volume_m3") and str(row.get("volume_m3")) not in ('', 'nan', 'None'):
                return f"V={float(row.get('volume_m3', 0) or 0):.2f} m³"
            return "—"
        inv_df["Sizing"] = inv_df.apply(size_summary, axis=1)
        inv_df_display = inv_df.rename(columns={"tag": "Tag", "type": "Equipment Type", "status": "Calculated?"})
        st.dataframe(
            inv_df_display[["Tag", "Equipment Type", "Sizing", "Calculated?"]],
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("No equipment items found.")

    st.markdown('<div class="section-header">🌊 Stream Summary</div>', unsafe_allow_html=True)
    stream_df = data.stream_summary()
    if not stream_df.empty:
        st.dataframe(stream_df, use_container_width=True, hide_index=True)
    else:
        st.info("No stream data found.")

    st.markdown('<div class="section-header">⚗️ Stream Compositions</div>', unsafe_allow_html=True)
    if not data.compositions.empty:
        comp_df = data.compositions.copy()
        # Ensure numeric columns are well-formatted if they exist
        for col in ["MoleFraction", "MassFraction"]:
            if col in comp_df.columns:
                comp_df[col] = pd.to_numeric(comp_df[col], errors="coerce")
                
        st.dataframe(comp_df, use_container_width=True, hide_index=True)
    else:
        st.info("No composition data found.")

    # Object type chart
    if inv:
        type_counts = pd.DataFrame(inv)["type"].value_counts().reset_index()
        type_counts.columns = ["Equipment Type", "Count"]
        fig = px.pie(
            type_counts, values="Count", names="Equipment Type",
            color_discrete_sequence=px.colors.sequential.Purpor,
            hole=0.45,
        )
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#c4b5fd", legend_font_color="#c4b5fd",
            margin=dict(t=10, b=10, l=10, r=10),
        )
        st.plotly_chart(fig, use_container_width=True)


# ============================================================
# Tab 2 – CAPEX Breakdown
# ============================================================
with tab2:
    st.markdown('<div class="section-header">💰 CAPEX — Per Equipment</div>', unsafe_allow_html=True)

    if report.equipment_costs:
        capex_rows = []
        for ec in report.equipment_costs:
            capex_rows.append({
                "Tag": ec.tag,
                "Type": ec.eq_type,
                "Size": ec.size_param,
                "Purchase Cost": ec.purchase_cost_current,
                "Bare-Module Cost": ec.bare_module_cost,
            })
        capex_df = pd.DataFrame(capex_rows)

        # Styled table with formatted costs
        disp = capex_df.copy()
        disp["Purchase Cost"] = disp["Purchase Cost"].apply(lambda v: fmt_usd(v))
        disp["Bare-Module Cost"] = disp["Bare-Module Cost"].apply(lambda v: fmt_usd(v))
        st.dataframe(disp, use_container_width=True, hide_index=True)

        # Horizontal bar chart – bare module costs
        fig_bar = px.bar(
            capex_df.sort_values("Bare-Module Cost"),
            x="Bare-Module Cost", y="Tag", orientation="h",
            color="Type",
            color_discrete_sequence=["#8b5cf6","#06b6d4","#f59e0b","#10b981","#ef4444","#6366f1"],
            labels={"Bare-Module Cost": "Bare-Module Cost (USD)", "Tag": "Equipment Tag"},
            title="Bare-Module Cost by Equipment",
        )
        fig_bar.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#c4b5fd", title_font_color="#c4b5fd",
            xaxis=dict(gridcolor="rgba(139,92,246,0.15)", zerolinecolor="rgba(139,92,246,0.3)"),
            yaxis=dict(gridcolor="rgba(0,0,0,0)"),
            legend_font_color="#c4b5fd",
            margin=dict(l=10, r=10, t=40, b=10),
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    st.markdown('<div class="section-header">📋 CAPEX Roll-Up Summary</div>', unsafe_allow_html=True)

    items = [
        ("Total Purchase Cost (C_P)",       report.total_purchase_cost,    "Equipment only, CEPCI-corrected"),
        ("Total Bare-Module Cost (C_BM)",    report.total_bare_module,      "Incl. installation factors F_BM"),
        ("Contingency",                      report.contingency,            f"{st.session_state.settings.contingency_pct:.0f}% of C_BM"),
        ("Engineering Fee",                  report.engineering_fee,        f"{st.session_state.settings.engineering_pct:.0f}% of C_BM"),
        ("ISBL (Inside Battery Limits)",     report.isbl,                   "C_BM + Contingency + Eng."),
        ("OSBL (Outside Battery Limits)",    report.osbl,                   "10% of ISBL"),
        ("Total Fixed Capital Investment",   report.total_fixed_capital,    "ISBL + OSBL"),
        ("Startup Cost",                     report.startup,                f"{st.session_state.settings.startup_pct:.0f}% of TFC"),
        ("Working Capital",                  report.working_capital,        f"{st.session_state.settings.working_capital_pct:.0f}% of TFC"),
        ("Total Project Cost",               report.total_project_cost,     "TFC + Startup + WC"),
    ]
    for label, val, note in items:
        c_l, c_v, c_n = st.columns([3, 1.5, 2])
        c_l.markdown(f"**{label}**")
        c_v.markdown(f"`{fmt_usd(val)}`")
        c_n.markdown(f"<span style='color:#7c7ca0;font-size:0.85em'>{note}</span>",
                     unsafe_allow_html=True)

    # Waterfall chart
    st.markdown('<div class="section-header">📈 CAPEX Waterfall</div>', unsafe_allow_html=True)
    wf_labels = ["Purchase Cost", "Installation\n(BM Factor)", "Contingency",
                  "Engineering", "OSBL", "Startup", "Working Capital"]
    wf_values = [
        report.total_purchase_cost,
        report.total_bare_module - report.total_purchase_cost,
        report.contingency,
        report.engineering_fee,
        report.osbl,
        report.startup,
        report.working_capital,
    ]
    fig_wf = go.Figure(go.Waterfall(
        name="CAPEX", orientation="v",
        x=wf_labels, y=wf_values,
        connector={"line": {"color": "rgba(139,92,246,0.3)"}},
        increasing={"marker": {"color": "#8b5cf6"}},
        decreasing={"marker": {"color": "#ef4444"}},
        totals={"marker": {"color": "#06b6d4"}},
        texttemplate="%{y:$.3s}",
        textposition="outside",
    ))
    fig_wf.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#c4b5fd", title_font_color="#c4b5fd",
        yaxis_title="USD",
        yaxis=dict(gridcolor="rgba(139,92,246,0.15)"),
        xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        showlegend=False,
        margin=dict(l=10, r=10, t=20, b=10),
    )
    st.plotly_chart(fig_wf, use_container_width=True)


# ============================================================
# Tab 3 – OPEX & Utilities
# ============================================================
with tab3:
    st.markdown('<div class="section-header">⚡ Utility OPEX Details</div>', unsafe_allow_html=True)

    if report.utility_rows:
        util_df = pd.DataFrame(report.utility_rows)
        util_disp = util_df.copy()
        util_disp["duty_kw"]         = util_disp["duty_kw"].apply(lambda v: f"{v:.2f} kW")
        util_disp["annual_cost_usd"] = util_disp["annual_cost_usd"].apply(fmt_usd)
        util_disp.columns = ["Tag", "Duty Type", "Utility", "Duty (kW)", "Annual Cost (USD/yr)"]
        st.dataframe(util_disp, use_container_width=True, hide_index=True)

        # Bar chart by utility
        util_group = util_df.groupby("utility")["annual_cost_usd"].sum().reset_index()
        fig_util = px.bar(
            util_group.sort_values("annual_cost_usd"),
            x="annual_cost_usd", y="utility", orientation="h",
            color="utility",
            color_discrete_sequence=["#06b6d4","#8b5cf6","#f59e0b","#10b981"],
            labels={"annual_cost_usd": "Annual Cost (USD/yr)", "utility": "Utility Type"},
            title="Annual Utility Cost by Type",
        )
        fig_util.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#c4b5fd", title_font_color="#c4b5fd",
            showlegend=False,
            xaxis=dict(gridcolor="rgba(139,92,246,0.15)"),
            yaxis=dict(gridcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=10, t=40, b=10),
        )
        st.plotly_chart(fig_util, use_container_width=True)
    else:
        st.info("No utility duties found in the flowsheet data.")

    st.markdown('<div class="section-header">🔍 Heat Exchanger Sizing & Cost Transparency</div>', unsafe_allow_html=True)
    
    # Exchanger list from report
    exchangers = [ec for ec in report.equipment_costs if ec.eq_type in ("Heat Exchanger", "Condenser", "Reboiler", "Kettle Reboiler")]
    if exchangers:
        for ec in exchangers:
            d = ec.details
            if d:
                duty_kw = d.get("duty_kw", 0.0)
                duty_str = f"{duty_kw/1000.0:.2f} MW" if duty_kw >= 1000.0 else f"{duty_kw:.1f} kW"
                ut_name = d.get("utility_name", "N/A")
                ut_temp = d.get("utility_temp_str", "N/A")
                u_val = d.get("u_value_w_m2k", 0.0)
                lmtd = d.get("lmtd_k", 0.0)
                area = d.get("area_m2", 0.0)
                n_shells = d.get("n_shells", 1)
                
                card_html = f"""
                <div style="background: rgba(255,255,255,0.03); border: 1px solid rgba(139,92,246,0.3); border-radius: 12px; padding: 1.2rem; margin-bottom: 1rem;">
                  <h4 style="color: #c4b5fd; margin-top: 0; margin-bottom: 0.8rem;">🔌 {ec.tag} ({ec.eq_type})</h4>
                  <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 1rem;">
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Duty</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{duty_str}</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Selected Utility</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{ut_name}</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Utility Temp</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{ut_temp}</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">U Value</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{u_val:.0f} W/m²K</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">LMTD</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{lmtd:.1f} K</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Calculated Area</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{area:.1f} m²</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">No. of Shells</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{n_shells}</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Purchase Cost</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{fmt_usd(ec.purchase_cost_current)}</strong>
                    </div>
                    <div>
                      <span style="color: #a0a0cc; font-size: 0.85em; text-transform: uppercase;">Bare-Module Cost</span><br>
                      <strong style="color: #fff; font-size: 1.1em;">{fmt_usd(ec.bare_module_cost)}</strong>
                    </div>
                  </div>
                </div>
                """
                st.markdown(card_html, unsafe_allow_html=True)
    else:
        st.info("No heat exchangers found to display sizing details.")

    st.markdown('<div class="section-header">⚗️ Raw Materials Database Calculator</div>', unsafe_allow_html=True)
    db_df = load_raw_material_db()
    if not db_df.empty:
        st.write("Loaded pricing from `dwsim_raw_material_costs.csv`.")
        feeds = detect_feed_streams(data)
        
        if "selected_feeds" not in st.session_state:
            st.session_state.selected_feeds = feeds
            
        selected_feeds = st.multiselect(
            "Select Feed Stream(s) for Raw Materials",
            options=data.streams[data.streams["ObjectType"] == "Material Stream"]["Tag"].tolist(),
            default=st.session_state.selected_feeds
        )
        st.session_state.selected_feeds = selected_feeds
        
        # Calculate table
        raw_df, calc_total = calculate_raw_materials(data, db_df, selected_feeds, s.operating_hours)
        
        if not raw_df.empty:
            if "custom_prices" not in st.session_state:
                st.session_state.custom_prices = {}
                
            # Apply custom overrides
            for idx, r_row in raw_df.iterrows():
                comp = r_row["Compound"]
                if comp in st.session_state.custom_prices:
                    raw_df.at[idx, "Price ($/kg)"] = st.session_state.custom_prices[comp]
                    raw_df.at[idx, "Annual Cost ($/yr)"] = r_row["Mass Flow (kg/yr)"] * st.session_state.custom_prices[comp]
            
            calc_total = raw_df["Annual Cost ($/yr)"].sum()
            
            edited_df = st.data_editor(
                raw_df,
                column_config={
                    "Feed Stream": st.column_config.TextColumn("Feed Stream", disabled=True),
                    "Compound": st.column_config.TextColumn("Compound", disabled=True),
                    "Mass Flow (kg/s)": st.column_config.NumberColumn("Mass Flow (kg/s)", format="%.4f", disabled=True),
                    "Mass Flow (kg/yr)": st.column_config.NumberColumn("Mass Flow (kg/yr)", format="%.0f", disabled=True),
                    "Price ($/kg)": st.column_config.NumberColumn("Price ($/kg)", min_value=0.0, format="%.4f"),
                    "Annual Cost ($/yr)": st.column_config.NumberColumn("Annual Cost ($/yr)", format="$%.0f", disabled=True),
                },
                use_container_width=True,
                hide_index=True,
                key="raw_materials_editor"
            )
            
            # Check for edits
            changes_made = False
            for idx, r_row in edited_df.iterrows():
                comp = r_row["Compound"]
                new_price = r_row["Price ($/kg)"]
                old_price = raw_df.at[idx, "Price ($/kg)"]
                if abs(new_price - old_price) > 1e-6:
                    st.session_state.custom_prices[comp] = new_price
                    changes_made = True
                    
            if changes_made:
                st.rerun()
                
            st.markdown(f"**Total Calculated Raw Material Cost:** `{fmt_usd(calc_total, '/yr')}`")
            
            # Sync with session state settings
            if abs(st.session_state.raw_material_cost - calc_total) > 1.0:
                st.session_state.raw_material_cost = calc_total
                st.rerun()
        else:
            st.info("No compounds or compositions found in the selected feed streams.")
    else:
        st.warning("Could not load database from `C:\\Users\\Dr. Shephali\\Downloads\\dwsim_raw_material_costs.csv`.")

    st.markdown('<div class="section-header">📋 OPEX Summary</div>', unsafe_allow_html=True)

    opex_items = [
        ("Raw Materials",          st.session_state.settings.raw_material_cost * (1.23 if st.session_state.settings.opex_method == "Turton (COMd)" else 1.0)),
        ("Utilities",              report.total_utility_opex),
        ("Labour",                 report.labour_opex),
        ("Maintenance",            report.maintenance_opex),
        ("Overhead",               report.overhead_opex),
        ("Insurance & Tax",        report.insurance_opex),
    ]
    for label, val in opex_items:
        c_l, c_v, c_bar = st.columns([2, 1, 3])
        c_l.markdown(f"**{label}**")
        c_v.markdown(f"`{fmt_usd(val, '/yr')}`")
        pct = val / report.total_opex * 100 if report.total_opex > 0 else 0
        c_bar.progress(min(int(pct), 100), text=f"{pct:.1f}%")

    st.markdown(f"#### 🔴 Total OPEX: **{fmt_usd(report.total_opex, '/yr')}**")

    # Utility cost reference table
    with st.expander("📘 Utility Cost Reference ($/GJ)", expanded=False):
        ref_df = pd.DataFrame(
            [{"Utility": k, "Cost ($/GJ)": v} for k, v in UTILITY_COSTS.items()]
        )
        st.dataframe(ref_df, use_container_width=True, hide_index=True)


# ============================================================
# Tab 4 – Dashboard
# ============================================================
with tab4:
    st.markdown('<div class="section-header">📊 Cost Summary Dashboard</div>', unsafe_allow_html=True)

    col_l, col_r = st.columns(2)

    with col_l:
        # CAPEX pie
        capex_labels = ["Purchase Cost", "Installation", "Contingency",
                         "Engineering", "OSBL", "Startup", "Working Capital"]
        capex_vals = [
            report.total_purchase_cost,
            report.total_bare_module - report.total_purchase_cost,
            report.contingency, report.engineering_fee,
            report.osbl, report.startup, report.working_capital,
        ]
        fig_pie_c = px.pie(
            names=capex_labels, values=capex_vals,
            title="CAPEX Breakdown",
            color_discrete_sequence=px.colors.sequential.Purpor[::-1],
            hole=0.4,
        )
        fig_pie_c.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#c4b5fd", title_font_color="#c4b5fd",
            legend_font_color="#c4b5fd",
            margin=dict(t=40, b=10, l=10, r=10),
        )
        st.plotly_chart(fig_pie_c, use_container_width=True)

    with col_r:
        # OPEX pie
        opex_labels = ["Raw Materials", "Utilities", "Labour", "Maintenance", "Overhead", "Insurance"]
        opex_vals   = [
            st.session_state.settings.raw_material_cost * (1.23 if st.session_state.settings.opex_method == "Turton (COMd)" else 1.0),
            report.total_utility_opex, report.labour_opex,
            report.maintenance_opex, report.overhead_opex, report.insurance_opex
        ]
        fig_pie_o = px.pie(
            names=opex_labels, values=opex_vals,
            title="OPEX Breakdown (Annual)",
            color_discrete_sequence=["#06b6d4","#8b5cf6","#f59e0b","#10b981"],
            hole=0.4,
        )
        fig_pie_o.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#c4b5fd", title_font_color="#c4b5fd",
            legend_font_color="#c4b5fd",
            margin=dict(t=40, b=10, l=10, r=10),
        )
        st.plotly_chart(fig_pie_o, use_container_width=True)

    # Annual cost breakdown gauge-style bar
    st.markdown('<div class="section-header">📅 Annual Cost Profile</div>', unsafe_allow_html=True)

    annual_items = {
        "Annualised CAPEX": report.annualised_capex,
        "Raw Materials":    st.session_state.settings.raw_material_cost * (1.23 if st.session_state.settings.opex_method == "Turton (COMd)" else 1.0),
        "Utilities":         report.total_utility_opex,
        "Labour":            report.labour_opex,
        "Maintenance":       report.maintenance_opex,
        "Overhead":          report.overhead_opex,
        "Insurance":         report.insurance_opex,
    }
    fig_annual = go.Figure()
    colors = ["#8b5cf6","#14b8a6","#06b6d4","#f59e0b","#10b981","#c084fc","#ef4444"]
    for i, (label, val) in enumerate(annual_items.items()):
        fig_annual.add_trace(go.Bar(
            name=label, x=[val], y=["Annual Cost"],
            orientation="h",
            marker_color=colors[i % len(colors)],
            text=[fmt_usd(val, "/yr")],
            textposition="inside",
        ))
    fig_annual.update_layout(
        barmode="stack",
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#c4b5fd", legend_font_color="#c4b5fd",
        xaxis_title="USD/year",
        xaxis=dict(gridcolor="rgba(139,92,246,0.15)"),
        yaxis=dict(gridcolor="rgba(0,0,0,0)"),
        height=200,
        margin=dict(l=10, r=10, t=20, b=10),
    )
    st.plotly_chart(fig_annual, use_container_width=True)

    # Profitability summary
    st.markdown('<div class="section-header">💡 Profitability Indicators</div>', unsafe_allow_html=True)
    p1, p2, p3 = st.columns(3)
    r_pct = st.session_state.settings.discount_rate_pct
    n_yrs = int(st.session_state.settings.plant_life_years)
    p1.metric("Annualised CAPEX", fmt_usd(report.annualised_capex, "/yr"),
              help=f"Capital Recovery Factor at {r_pct}% over {n_yrs} years")
    p2.metric("Total Annual Cost", fmt_usd(report.total_annual_cost, "/yr"))
    p3.metric("Simple Payback Period",
              "N/A" if report.total_opex <= 0 else
              f"{report.total_fixed_capital/report.total_opex:.1f} yr",
              help="Fixed capital / Annual OPEX")


# ============================================================
# Tab 5 – Export
# ============================================================
with tab5:
    st.markdown('<div class="section-header">📄 Export Cost Report</div>', unsafe_allow_html=True)
    st.markdown("Download a multi-sheet Excel workbook containing all cost tables.")

    engine_export = CostEngine(data, st.session_state.settings)
    dfs = engine_export.to_dataframes(report)

    # Build Excel in memory
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # 1. Write the Settings sheet first so VLOOKUPs evaluate nicely
        settings_data = {
            "Parameter": [
                "CEPCI", "Material Factor", "Operating Hours/yr",
                "Electricity Price ($/kWh)", "Steam Utility", "Cooling Utility",
                "Contingency %", "Engineering %", "Plant Life (yr)",
                "Discount Rate %", "Labour Cost ($/yr)", "Maintenance % of FCI",
                "Insurance & Tax % of FCI", "Startup Cost %", "Working Capital %",
                "Raw Material Cost ($/yr)", "Location Factor", "Currency Conversion",
                "Currency Symbol", "CAPEX Mode", "Overhead % of Labour + Maint",
                "Pump Default Material", "HX Default Material", "Vessel Default Material",
                "Compressor Default Material", "Tray Default Material", "Tray Type"
            ],
            "Value": [
                st.session_state.settings.cepci,
                st.session_state.settings.material_factor,
                st.session_state.settings.operating_hours,
                st.session_state.settings.elec_price_per_kwh,
                st.session_state.settings.steam_utility,
                st.session_state.settings.cooling_utility,
                st.session_state.settings.contingency_pct,
                st.session_state.settings.engineering_pct,
                st.session_state.settings.plant_life_years,
                st.session_state.settings.discount_rate_pct,
                report.labour_opex / 2.73 if st.session_state.settings.opex_method == "Turton (COMd)" else report.labour_opex,
                st.session_state.settings.maintenance_pct,
                st.session_state.settings.insurance_pct,
                st.session_state.settings.startup_pct,
                st.session_state.settings.working_capital_pct,
                st.session_state.settings.raw_material_cost,
                st.session_state.settings.location_factor,
                st.session_state.settings.currency_conversion,
                st.session_state.settings.currency_symbol,
                st.session_state.settings.capex_mode,
                st.session_state.settings.overhead_pct,
                st.session_state.settings.pump_material,
                st.session_state.settings.hx_material,
                st.session_state.settings.vessel_material,
                st.session_state.settings.compressor_material,
                st.session_state.settings.tray_material,
                st.session_state.settings.tray_type
            ],
        }
        pd.DataFrame(settings_data).to_excel(writer, sheet_name="Settings", index=False)
        
        # Add the utility pricing lookup table to Settings sheet columns D and E
        ws_sett = writer.sheets["Settings"]
        ws_sett["D1"] = "Utility Name"
        ws_sett["E1"] = "Price ($/GJ)"
        ws_sett["D1"].font = Font(color="FFFFFF", bold=True)
        ws_sett["E1"].font = Font(color="FFFFFF", bold=True)
        ws_sett["D1"].fill = PatternFill("solid", fgColor="4F2AF5")
        ws_sett["E1"].fill = PatternFill("solid", fgColor="4F2AF5")
        
        # Prices scale with currency conversion settings factor B19
        util_prices = [
            ("LP Steam (3 bar)", "=6.62 * B19"),
            ("MP Steam (10 bar)", "=8.22 * B19"),
            ("HP Steam (42 bar)", "=12.32 * B19"),
            ("Cooling Water", "=0.354 * B19"),
            ("Chilled Water", "=4.77 * B19"),
            ("Refrigerant (-25°C)", "=23.53 * B19"),
            ("Electricity", "=B5*277.78*B19"),
            ("Fuel (natural gas)", "=4.47 * B19"),
        ]
        for idx, (name, val) in enumerate(util_prices, start=2):
            ws_sett[f"D{idx}"] = name
            ws_sett[f"E{idx}"] = val

        # 2. Write the rest of the dataframes
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 3. Apply styles and override cell values with formulas where necessary
        all_sheet_names = list(writer.sheets.keys())
        for sheet_name in all_sheet_names:
            if sheet_name == "Settings":
                ws = writer.sheets[sheet_name]
                ws.column_dimensions["A"].width = 30
                ws.column_dimensions["B"].width = 30
                ws.column_dimensions["D"].width = 25
                ws.column_dimensions["E"].width = 20
                continue
                
            ws = writer.sheets[sheet_name]
            
            # Style header row
            fill = PatternFill("solid", fgColor="4F2AF5")
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.freeze_panes = "A2"
            
            # Apply formulas depending on the sheet
            if sheet_name == "Equipment_CAPEX":
                for row in range(2, ws.max_row + 1):
                    # We fetch the original list item to calculate fbm_factor
                    if row - 2 < len(report.equipment_costs):
                        ec = report.equipment_costs[row - 2]
                        # Determine fbm_factor from scaled or unscaled values
                        fbm_factor = ec.bare_module_cost / ec.purchase_cost_current if ec.purchase_cost_current > 0 else 1.0
                    else:
                        fbm_factor = 3.0
                    ws[f"E{row}"] = f"=D{row} * (Settings!$B$2 / 397) * Settings!$B$18 * Settings!$B$19"
                    ws[f"F{row}"] = f"=E{row} * {fbm_factor:.3f}"
                    
            elif sheet_name == "Utility_OPEX":
                # Column E is Price ($/GJ), Column F is Annual Cost (USD/yr)
                for row in range(2, ws.max_row + 1):
                    ws[f"E{row}"] = f'=VLOOKUP(C{row}, Settings!$D$2:$E$9, 2, FALSE)'
                    ws[f"F{row}"] = f'=D{row} * 3600 * Settings!$B$4 / 1000000 * E{row}'
                    
            elif sheet_name == "CAPEX_Summary":
                ws["B2"] = "=SUM(Equipment_CAPEX!E:E)"
                ws["B3"] = "=SUM(Equipment_CAPEX!F:F)"
                
                if st.session_state.settings.capex_mode == "Rigorous (System Expansion)":
                    ws["B4"] = "=B3 * 0.15"  # contingency
                    ws["B5"] = "=B3 * 0.03"  # engineering
                    ws["B6"] = "=1.18 * B3"  # ISBL is C_TM
                    ws["B7"] = "=0"          # OSBL is 0
                    ws["B8"] = "=B6 + 0.35 * B2"  # FCI
                else:
                    ws["B4"] = "=B3 * Settings!$B$8 / 100"
                    ws["B5"] = "=B3 * Settings!$B$9 / 100"
                    ws["B6"] = "=B3 + B4 + B5"
                    ws["B7"] = "=B6 * 0.10"
                    ws["B8"] = "=B6 + B7"
                    
                ws["B9"] = "=B8 * Settings!$B$15 / 100"
                ws["B10"] = "=B8 * Settings!$B$16 / 100"
                ws["B11"] = "=B8 + B9 + B10"
                
            elif sheet_name == "OPEX_Summary":
                if st.session_state.settings.opex_method == "Turton (COMd)":
                    ws["B2"] = "=Settings!$B$17 * 1.23"  # Raw materials
                    ws["B3"] = "=SUM(Utility_OPEX!F:F) * 1.23" # Utilities
                    ws["B4"] = "=Settings!$B$12 * 2.73"  # Labour
                    ws["B5"] = "=0.06 * CAPEX_Summary!$B$8" # Maintenance is 6% of FCI
                    ws["B6"] = "=0.60 * (B4 + B5)"  # Overhead is 60% of (Labour + Maintenance)
                    ws["B7"] = "=0.03 * CAPEX_Summary!$B$8" # Insurance & Tax is 3% of FCI
                else:
                    ws["B2"] = "=Settings!$B$17"  # Raw materials
                    ws["B3"] = "=SUM(Utility_OPEX!F:F)" # Utilities
                    ws["B4"] = "=Settings!$B$12"  # Labour
                    ws["B5"] = "=Settings!$B$13 / 100 * CAPEX_Summary!$B$8" # Maintenance
                    ws["B6"] = "=Settings!$B$22 / 100 * (B4 + B5)"  # Overhead
                    ws["B7"] = "=Settings!$B$14 / 100 * CAPEX_Summary!$B$8" # Insurance
                    
                ws["B8"] = "=SUM(B2:B7)"
                ws["B9"] = "=IF(Settings!$B$11>0, CAPEX_Summary!$B$8 * (Settings!$B$11/100 * (1 + Settings!$B$11/100)^Settings!$B$10) / ((1 + Settings!$B$11/100)^Settings!$B$10 - 1), CAPEX_Summary!$B$8 / Settings!$B$10)"
                ws["B10"] = "=B8 + B9"
            
            # Auto column width
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    buf.seek(0)

    flowsheet_stem = Path(data.flowsheet_name).stem
    st.download_button(
        label="⬇️ Download Cost Report (.xlsx)",
        data=buf,
        file_name=f"{flowsheet_stem}_cost_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("### Included Sheets")
    sheets_info = [
        ("Equipment_CAPEX",  "Per-equipment purchase and bare-module costs"),
        ("Utility_OPEX",     "Per-equipment annual utility costs"),
        ("CAPEX_Summary",    "Rolled-up CAPEX from purchase to total project cost"),
        ("OPEX_Summary",     "Annual OPEX breakdown with annualised CAPEX"),
        ("Settings",         "All settings/assumptions used for this estimate"),
    ]
    for sname, sdesc in sheets_info:
        st.markdown(f"- **{sname}** — {sdesc}")

    # Show raw data preview
    with st.expander("🔍 Preview Export Data", expanded=False):
        for name, df in dfs.items():
            st.markdown(f"**{name}**")
            st.dataframe(df.head(10), use_container_width=True, hide_index=True)

# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------
st.markdown("---")
st.markdown(
    "<div style='text-align:center;color:#4a4a6a;font-size:0.82rem'>"
    "DWSIM Cost Estimator · Built with Streamlit · Turton et al. (2012) correlations · "
    "For educational &amp; pre-feasibility purposes only"
    "</div>",
    unsafe_allow_html=True,
)
