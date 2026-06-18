"""
excel_parser.py
Parses the DWSIM costing introspection Excel workbook produced by
costing_engine_variables.py into clean DataFrames for the cost estimator.
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any, Union

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_sheet(wb: openpyxl.Workbook, name: str) -> pd.DataFrame:
    """Return a DataFrame for a sheet, or an empty one if missing/no data."""
    if name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[name]
    data = list(ws.iter_rows(values_only=True))
    if not data or not any(v is not None for v in data[0]):
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    rows = [list(r) for r in data[1:] if any(v is not None for v in r)]
    return pd.DataFrame(rows, columns=headers)


def _to_float(val: Any) -> float | None:
    """Safely convert a cell value to float."""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _first_number(text: str | None) -> float | None:
    """Extract the first numeric value from a pipe-separated string."""
    if not text:
        return None
    parts = str(text).split("|")
    for p in parts:
        v = _to_float(p.strip())
        if v is not None:
            return v
    return None


def _pipe_to_list(text: str | None) -> list[float]:
    """Parse a pipe-separated string of numbers into a list."""
    if not text:
        return []
    result = []
    for p in str(text).split("|"):
        v = _to_float(p.strip())
        if v is not None:
            result.append(v)
    return result


# ---------------------------------------------------------------------------
# Public parser
# ---------------------------------------------------------------------------

class DWSIMData:
    """Container for all parsed tables from one introspection Excel."""

    def __init__(self, path: Union[str, Path, io.BytesIO]):
        # Accept both file paths and in-memory BytesIO (from Streamlit uploader)
        if isinstance(path, (str, Path)):
            self.path = Path(path)
            source = str(self.path)
        else:
            # BytesIO or any file-like object
            self.path = Path(getattr(path, 'name', 'uploaded.xlsx'))
            path.seek(0)
            source = path
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

        self.summary = self._parse_summary(wb)
        self.objects = _read_sheet(wb, "Objects")
        self.streams = _read_sheet(wb, "Streams")
        self.compositions = _read_sheet(wb, "Compositions")
        self.energy_streams = _read_sheet(wb, "EnergyStreams")
        self.distillation_columns = _read_sheet(wb, "DistillationColumns")
        self.column_stage_profiles = _read_sheet(wb, "ColumnStageProfiles")
        self.heat_exchangers = _read_sheet(wb, "HeatExchangers")
        self.pumps = _read_sheet(wb, "Pumps")
        self.compressors = _read_sheet(wb, "Compressors")
        self.reactors = _read_sheet(wb, "Reactors")
        self.other_equipment = _read_sheet(wb, "OtherEquipment")
        self.connections = _read_sheet(wb, "Connections")
        self.costing_vars = _read_sheet(wb, "CostingVariables")
        wb.close()

        # Build a quick lookup: tag -> {variable: value}
        self.costing_lookup: dict[str, dict[str, str]] = {}
        if not self.costing_vars.empty:
            for _, row in self.costing_vars.iterrows():
                tag = str(row.get("Tag", "")).strip()
                var = str(row.get("Variable", "")).strip()
                val = row.get("Value", "")
                if tag not in self.costing_lookup:
                    self.costing_lookup[tag] = {}
                self.costing_lookup[tag][var] = val

    # ------------------------------------------------------------------
    def _parse_summary(self, wb) -> dict:
        if "Summary" not in wb.sheetnames:
            return {}
        ws = wb["Summary"]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                result[str(row[0])] = row[1]
        return result

    # ------------------------------------------------------------------
    def get_costing_value(self, tag: str, *keys: str) -> float | None:
        """Return the first numeric costing variable matching any of `keys` for `tag`."""
        lookup = self.costing_lookup.get(str(tag).strip(), {})
        for key in keys:
            val = lookup.get(key)
            if val is not None:
                f = _to_float(val)
                if f is not None:
                    return f
        return None

    # ------------------------------------------------------------------
    def equipment_inventory(self) -> list[dict]:
        """
        Returns a unified list of equipment dicts, each with normalised fields
        needed by the cost engine.
        """
        items: list[dict] = []

        # --- Distillation Columns ---
        if not self.distillation_columns.empty and \
                "No data" not in str(self.distillation_columns.columns.tolist()):
            for _, row in self.distillation_columns.iterrows():
                tag = str(row.get("Tag", "")).strip()
                if not tag:
                    continue
                n_stages = _to_float(row.get("NumberOfStages")) or \
                           self.get_costing_value(tag, "NumberOfStages", "NStages")
                diam = _to_float(row.get("EstimatedDiameter_m")) or \
                       self.get_costing_value(tag, "EstimatedDiameter", "Diameter")
                height = _to_float(row.get("EstimatedHeight_m")) or \
                         self.get_costing_value(tag, "EstimatedHeight", "Height")
                tray_spacing = _to_float(row.get("TraySpacing_m")) or 0.6
                cond_duty = _to_float(row.get("CondenserDuty_kW"))
                reb_duty = abs(_to_float(row.get("ReboilerDuty_kW")) or 0)
                cond_duty_abs = abs(cond_duty) if cond_duty is not None else 0.0

                items.append({
                    "tag": tag,
                    "type": "Distillation Column",
                    "n_stages": n_stages,
                    "diameter_m": diam,
                    "height_m": height,
                    "tray_spacing_m": tray_spacing,
                    "condenser_duty_kw": cond_duty_abs,
                    "reboiler_duty_kw": reb_duty,
                    "reflux_ratio": _to_float(row.get("RefluxRatio")),
                    "status": str(row.get("CalculationStatus", "")),
                })

        # --- Heat Exchangers ---
        if not self.heat_exchangers.empty and \
                "No data" not in str(self.heat_exchangers.columns.tolist()):
            for _, row in self.heat_exchangers.iterrows():
                tag = str(row.get("Tag", "")).strip()
                if not tag:
                    continue
                area = _to_float(row.get("Area_m2")) or \
                       self.get_costing_value(tag, "Area", "HeatTransferArea")
                duty_kw = abs(_to_float(row.get("Duty_kW")) or
                              self.get_costing_value(tag, "Duty", "HeatDuty") or 0)
                items.append({
                    "tag": tag,
                    "type": "Heat Exchanger",
                    "area_m2": area,
                    "duty_kw": duty_kw,
                    "status": str(row.get("CalculationStatus", "")),
                })

        # --- Pumps ---
        if not self.pumps.empty and \
                "No data" not in str(self.pumps.columns.tolist()):
            for _, row in self.pumps.iterrows():
                tag = str(row.get("Tag", "")).strip()
                if not tag:
                    continue
                power_kw = _to_float(row.get("Power_kW")) or \
                               self.get_costing_value(tag, "Power", "ActualPower")
                if power_kw is not None:
                    power_kw = abs(power_kw)
                items.append({
                    "tag": tag,
                    "type": "Pump",
                    "power_kw": power_kw,
                    "status": str(row.get("CalculationStatus", "")),
                })

        # --- Compressors ---
        if not self.compressors.empty and \
                "No data" not in str(self.compressors.columns.tolist()):
            for _, row in self.compressors.iterrows():
                tag = str(row.get("Tag", "")).strip()
                if not tag:
                    continue
                power_kw = _to_float(row.get("Power_kW")) or \
                               self.get_costing_value(tag, "Power", "ActualPower")
                if power_kw is not None:
                    power_kw = abs(power_kw)
                items.append({
                    "tag": tag,
                    "type": "Compressor",
                    "power_kw": power_kw,
                    "status": str(row.get("CalculationStatus", "")),
                })

        # --- Reactors ---
        if not self.reactors.empty and \
                "No data" not in str(self.reactors.columns.tolist()):
            for _, row in self.reactors.iterrows():
                tag = str(row.get("Tag", "")).strip()
                if not tag:
                    continue
                volume = _to_float(row.get("Volume")) or \
                         self.get_costing_value(tag, "Volume", "VesselVolume")
                items.append({
                    "tag": tag,
                    "type": "Reactor",
                    "volume_m3": volume,
                    "status": str(row.get("CalculationStatus", "")),
                })

        # --- Other Equipment (Mixers, Splitters, Valves, etc.) ---
        if not self.other_equipment.empty and \
                "No data" not in str(self.other_equipment.columns.tolist()):
            for _, row in self.other_equipment.iterrows():
                tag = str(row.get("Tag", "")).strip()
                obj_type = str(row.get("ObjectType", "Other")).strip()
                if not tag:
                    continue
                items.append({
                    "tag": tag,
                    "type": obj_type,
                    "status": str(row.get("CalculationStatus", "")),
                })

        # Fallback: if nothing above produced items, use Objects sheet
        if not items and not self.objects.empty:
            for _, row in self.objects.iterrows():
                tag = str(row.get("Tag", "")).strip()
                obj_type = str(row.get("ObjectType", "Unknown")).strip()
                if not tag or obj_type in ("Material Stream", "Energy Stream"):
                    continue
                items.append({
                    "tag": tag,
                    "type": obj_type,
                    "status": str(row.get("CalculationStatus", "")),
                })

        return items

    # ------------------------------------------------------------------
    def stream_summary(self) -> pd.DataFrame:
        """Return a clean stream summary DataFrame."""
        if self.streams.empty:
            return pd.DataFrame()
        cols_of_interest = [
            "Tag", "ObjectType", "Temperature_C", "Pressure_bar",
            "MassFlow_kg_s", "MolarFlow_mol_s", "VaporFraction",
        ]
        existing = [c for c in cols_of_interest if c in self.streams.columns]
        return self.streams[existing].copy()

    # ------------------------------------------------------------------
    @property
    def flowsheet_name(self) -> str:
        fp = self.summary.get("Flowsheet", "")
        if fp:
            try:
                return Path(str(fp)).name
            except Exception:
                pass
        return self.path.name
