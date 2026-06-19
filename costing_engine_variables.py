"""
DWSIM generic introspection and costing-variable extractor.

Usage:
    python costing_engine_variables.py "C:\\path\\flowsheet.dwxmz"
    python costing_engine_variables.py "C:\\path\\flowsheet.dwxmz" --output out.xlsx
    python costing_engine_variables.py "C:\\path\\flowsheet.dwxmz" --dwsim "C:\\Users\\...\\AppData\\Local\\DWSIM"

The extractor is intentionally reflection-heavy. DWSIM's report writer often reads
from nested result objects, private backing fields, and collection items rather
than only first-level GUI properties. This script therefore exports:
    - public root properties, fields, and methods,
    - recursive public + non-public discovery rows,
    - targeted process/costing rows for common equipment classes,
    - a filtered CAPEX/OPEX costing-variable table.
"""

from __future__ import annotations

import argparse
import math
import os
import re
import sys
import traceback
from collections import OrderedDict, defaultdict
from datetime import datetime
from typing import Any, Iterable


MAX_CELL_CHARS = 32000
MAX_DISCOVERY_DEPTH = 4
MAX_COLLECTION_ITEMS = 300

REQUIRED_SHEETS = [
    "Objects",
    "Properties",
    "Fields",
    "Streams",
    "Compositions",
    "EnergyStreams",
    "DistillationColumns",
    "HeatExchangers",
    "Pumps",
    "Compressors",
    "Connections",
    "CostingVariables",
    "DiscoveryDump",
]


TYPE_MAP = [
    ("MaterialStream", "Material Stream"),
    ("EnergyStream", "Energy Stream"),
    ("DistillationColumn", "Distillation Column"),
    ("AbsorptionColumn", "Absorption Column"),
    ("ShortcutColumn", "Shortcut Column"),
    ("HeatExchanger", "Heat Exchanger"),
    ("Heater", "Heat Exchanger"),
    ("Cooler", "Heat Exchanger"),
    ("Pump", "Pump"),
    ("Compressor", "Compressor"),
    ("Expander", "Compressor / Expander"),
    ("Valve", "Valve"),
    ("Mixer", "Mixer"),
    ("Splitter", "Splitter"),
    ("Reactor_Conversion", "Reactor"),
    ("Reactor_CSTR", "Reactor"),
    ("Reactor_PFR", "Reactor"),
    ("Reactor_Equilibrium", "Reactor"),
    ("Reactor_Gibbs", "Reactor"),
    ("Reactor", "Reactor"),
    ("Tank", "Vessel"),
    ("Vessel", "Vessel"),
    ("PipeSegment", "Pipe"),
    ("Pipe", "Pipe"),
    ("CustomUO", "Custom Unit Operation"),
    ("CapeOpen", "Custom Unit Operation"),
]

COSTING_KEYWORDS = {
    "area", "diameter", "height", "volume", "length", "weight", "mass",
    "duty", "power", "energy", "work", "heat",
    "stage", "stages", "tray", "trays", "plate", "plates",
    "spacing", "efficiency", "eff", "reflux", "boilup",
    "pressure", "temperature", "flow", "massflow", "molarflow",
    "volumetric", "density", "viscosity",
    "transfer", "coefficient", "uvalue", "overallu", "lmtd", "ntu",
    "tube", "shell", "baffle", "pass", "fouling",
    "npsh", "head", "ratio", "conversion", "residence",
    "mechanical", "sizing", "design", "wall", "thickness",
}

CAPEX_HINTS = {
    "area", "diameter", "height", "volume", "length", "weight", "mass",
    "stage", "tray", "plate", "spacing", "tube", "shell", "baffle",
    "pass", "mechanical", "sizing", "design", "wall", "thickness",
}

RECURSE_NAME_HINTS = {
    "result", "results", "profile", "profiles", "stage", "stages", "tray", "trays",
    "plate", "plates", "sizing", "design", "mechanical", "geometry",
    "condenser", "reboiler", "column", "section",
    "phase", "phases", "compound", "compounds", "properties", "property",
    "parameter", "parameters", "spec", "specification", "specifications",
    "efficiency", "efficiencies", "pressure", "temperature", "flow", "flows",
    "duty", "heat", "power", "energy", "area", "diameter", "height", "volume",
    "tc", "pc", "lc", "vc", "_tc", "_pc", "_lc", "_vc",
}

NO_RECURSE_NAME_HINTS = {
    "graphicobject", "flowsheet", "simulation", "simulator", "host", "form",
    "editor", "surface", "canvas", "document", "spreadsheet", "script",
    "chart", "annotation", "settings", "mobile", "localization", "unitssystem",
    "propertypackage", "propertypackages", "reactionsets", "reactions",
    "file", "filepath", "image", "icon", "font", "color",
}


# These globals are filled after pythonnet/DWSIM assemblies are loaded.
System = None
BindingFlags = None
RuntimeHelpers = None
BIND_PUBLIC = None
BIND_ALL = None


def safe(fn, default=None):
    try:
        value = fn()
        return default if value is None else value
    except Exception:
        return default


def is_missing(value: Any) -> bool:
    return value is None or value == ""


def is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        f = float(value)
        return not math.isnan(f) and not math.isinf(f)
    except Exception:
        return False


def round_float(value: Any, digits: int = 8):
    if not is_number(value):
        return ""
    f = float(value)
    if abs(f) >= 1e15:
        return str(f)
    return round(f, digits)


def k_to_c(value: Any):
    return round_float(float(value) - 273.15, 5) if is_number(value) else ""


def pa_to_bar(value: Any):
    return round_float(float(value) / 1e5, 8) if is_number(value) else ""


def watt_to_kw(value: Any):
    return round_float(float(value) / 1000.0, 8) if is_number(value) else ""


def clean_text(value: Any, limit: int = MAX_CELL_CHARS) -> str:
    if value is None:
        return ""
    text = str(value)
    import re
    # Remove illegal characters for Excel (openpyxl)
    text = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', text)
    return text[:limit]


def to_excel(value: Any):
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, str)):
        if isinstance(value, float):
            return round_float(value)
        return clean_text(value) if isinstance(value, str) else value
    try:
        type_name = value.GetType().FullName
        if type_name == "System.String":
            return clean_text(value)
        if value.GetType().IsEnum:
            return clean_text(value)
    except Exception:
        pass
    if is_number(value):
        return round_float(value)
    return clean_text(value)


def sanitize_name(value: Any, fallback: str = "") -> str:
    text = clean_text(value, 160).strip()
    if not text:
        return fallback
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^0-9A-Za-z_.:-]+", "_", text)
    return text.strip("_") or fallback


def sequence_to_text(value: Any, limit: int = 80) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return " | ".join(clean_text(x, 160) for x in value[:limit])
    return clean_text(value, 1000)


def get_type_name(obj: Any) -> str:
    if obj is None:
        return ""
    return safe(lambda: obj.GetType().FullName, type(obj).__name__)


def object_hash(obj: Any) -> str:
    if obj is None:
        return "None"
    if RuntimeHelpers is not None:
        return f"{get_type_name(obj)}:{safe(lambda: RuntimeHelpers.GetHashCode(obj), id(obj))}"
    return f"{get_type_name(obj)}:{id(obj)}"


def is_scalar(value: Any) -> bool:
    if value is None or isinstance(value, (str, bool, int, float)):
        return True
    try:
        t = value.GetType()
        return bool(t.IsPrimitive or t.IsEnum or t.FullName in {
            "System.String",
            "System.Decimal",
            "System.DateTime",
            "System.Guid",
            "System.TimeSpan",
        })
    except Exception:
        return False


def is_iterable_net(value: Any) -> bool:
    if value is None or is_scalar(value):
        return False
    try:
        from System.Collections import IEnumerable
        return isinstance(value, IEnumerable)
    except Exception:
        return hasattr(value, "__iter__")


def net_count(value: Any):
    for attr in ("Count", "Length"):
        result = safe(lambda a=attr: getattr(value, a), None)
        if result is not None:
            return to_excel(result)
    return ""


def iter_net_items(value: Any, limit: int = MAX_COLLECTION_ITEMS):
    if value is None or is_scalar(value):
        return

    # IDictionary-like objects: emit keyed items if possible.
    keys = safe(lambda: value.Keys, None)
    if keys is not None:
        count = 0
        for key in keys:
            if count >= limit:
                return
            item = safe(lambda k=key: value[k], None)
            yield clean_text(key, 120), item
            count += 1
        return

    count = 0
    try:
        for item in value:
            if count >= limit:
                return
            key = safe(lambda i=item: i.Key, None)
            val = safe(lambda i=item: i.Value, None)
            if key is not None and val is not None:
                yield clean_text(key, 120), val
            else:
                yield str(count), item
            count += 1
    except Exception:
        return


def method_signature(method) -> str:
    params = []
    for p in safe(lambda: method.GetParameters(), []) or []:
        params.append(f"{safe(lambda p=p: p.ParameterType.Name, '?')} {safe(lambda p=p: p.Name, '?')}")
    ret = safe(lambda: method.ReturnType.Name, "?")
    return f"{ret} {safe(lambda: method.Name, '?')}({', '.join(params)})"


def should_recurse_member(name: str, path: str, value: Any, depth: int) -> bool:
    if value is None or is_scalar(value):
        return False
    text = f"{name} {path} {get_type_name(value)}".lower()
    compact = re.sub(r"[^a-z0-9_]+", "", text)
    if any(hint in compact for hint in NO_RECURSE_NAME_HINTS):
        return False
    if depth == 0:
        return any(hint in compact for hint in RECURSE_NAME_HINTS)
    return any(hint in compact for hint in RECURSE_NAME_HINTS)


def classify(obj: Any) -> str:
    class_name = type(obj).__name__
    full_name = get_type_name(obj)
    haystack = f"{class_name} {full_name}".lower()
    for keyword, label in TYPE_MAP:
        if keyword.lower() in haystack:
            return label
    return "Custom Unit Operation" if "custom" in haystack else f"Other ({class_name})"


def get_tag(obj: Any, fallback: str = "") -> str:
    candidates = [
        lambda: obj.GraphicObject.Tag,
        lambda: obj.GraphicObject.Name,
        lambda: obj.Tag,
        lambda: obj.Name,
    ]
    for fn in candidates:
        value = safe(fn, None)
        if value:
            return clean_text(value, 200)
    return clean_text(fallback, 200)


def get_description(obj: Any) -> str:
    for attr in ("Description", "Annotation", "Comments", "Comment"):
        value = safe(lambda a=attr: getattr(obj, a), None)
        if value:
            return clean_text(value, 1000)
    return ""


def get_calc_status(obj: Any) -> str:
    for attr in ("Calculated", "CalculationStatus", "Status", "IsCalculated"):
        value = safe(lambda a=attr: getattr(obj, a), None)
        if value is not None:
            return clean_text(value, 200)
    return ""


def get_error_message(obj: Any) -> str:
    for attr in ("ErrorMessage", "Error", "LastError", "ExceptionMessage"):
        value = safe(lambda a=attr: getattr(obj, a), None)
        if value:
            return clean_text(value, 1000)
    return ""


def member_rows(obj: Any, tag: str, category: str, class_name: str):
    properties, fields, methods = [], [], []
    net_type = safe(lambda: obj.GetType(), None)
    if net_type is None:
        return properties, fields, methods

    for prop in safe(lambda: net_type.GetProperties(BIND_PUBLIC), []) or []:
        if safe(lambda p=prop: p.GetIndexParameters().Length, 1) != 0:
            continue
        value, error = "", ""
        try:
            if prop.CanRead:
                raw = prop.GetValue(obj, None)
                value = to_excel(raw) if is_scalar(raw) else ""
        except Exception as exc:
            error = clean_text(exc, 300)
        properties.append({
            "Tag": tag,
            "ObjectType": category,
            "ClassName": class_name,
            "Name": prop.Name,
            "DeclaringType": safe(lambda p=prop: p.DeclaringType.FullName, ""),
            "PropertyType": safe(lambda p=prop: p.PropertyType.FullName, ""),
            "CanRead": safe(lambda p=prop: p.CanRead, ""),
            "CanWrite": safe(lambda p=prop: p.CanWrite, ""),
            "Value": value,
            "Error": error,
        })

    for field in safe(lambda: net_type.GetFields(BIND_PUBLIC), []) or []:
        value, error = "", ""
        try:
            raw = field.GetValue(obj)
            value = to_excel(raw) if is_scalar(raw) else ""
        except Exception as exc:
            error = clean_text(exc, 300)
        fields.append({
            "Tag": tag,
            "ObjectType": category,
            "ClassName": class_name,
            "Name": field.Name,
            "DeclaringType": safe(lambda f=field: f.DeclaringType.FullName, ""),
            "FieldType": safe(lambda f=field: f.FieldType.FullName, ""),
            "IsStatic": safe(lambda f=field: f.IsStatic, ""),
            "Value": value,
            "Error": error,
        })

    for method in safe(lambda: net_type.GetMethods(BIND_PUBLIC), []) or []:
        if safe(lambda m=method: m.IsSpecialName, False):
            continue
        methods.append({
            "Tag": tag,
            "ObjectType": category,
            "ClassName": class_name,
            "Name": safe(lambda m=method: m.Name, ""),
            "Signature": method_signature(method),
            "DeclaringType": safe(lambda m=method: m.DeclaringType.FullName, ""),
            "ReturnType": safe(lambda m=method: m.ReturnType.FullName, ""),
            "IsStatic": safe(lambda m=method: m.IsStatic, ""),
        })

    return properties, fields, methods


def discovery_rows(
    obj: Any,
    tag: str,
    category: str,
    root_name: str = "",
    max_depth: int = MAX_DISCOVERY_DEPTH,
):
    rows = []
    visited = set()

    def add_row(path, kind, member_type, value_type="", value="", access="", error=""):
        rows.append({
            "Tag": tag,
            "ObjectType": category,
            "Path": clean_text(path, 2000),
            "MemberKind": kind,
            "MemberType": clean_text(member_type, 1000),
            "ValueType": clean_text(value_type, 1000),
            "Value": to_excel(value),
            "Access": access,
            "Depth": path.count(".") + path.count("["),
            "IsCostingVariable": "YES" if is_costing_key(path) else "",
            "Error": clean_text(error, 500),
        })

    def walk(current, prefix, depth):
        if current is None or depth > max_depth:
            return
        if is_scalar(current):
            add_row(prefix, "Value", "", type(current).__name__, current, "", "")
            return

        oid = object_hash(current)
        if oid in visited:
            return
        visited.add(oid)

        net_type = safe(lambda: current.GetType(), None)
        if net_type is None:
            add_row(prefix, "Object", type(current).__name__, "", clean_text(current), "", "")
            return

        if is_iterable_net(current):
            add_row(prefix, "Collection", safe(lambda: net_type.FullName, ""), "", f"Count={net_count(current)}")
            if depth < max_depth:
                for key, item in iter_net_items(current):
                    walk(item, f"{prefix}[{sanitize_name(key, 'item')}]", depth + 1)
            return

        for prop in safe(lambda: net_type.GetProperties(BIND_ALL), []) or []:
            if safe(lambda p=prop: p.GetIndexParameters().Length, 1) != 0:
                continue
            name = safe(lambda p=prop: p.Name, "")
            if not name:
                continue
            path = f"{prefix}.{name}" if prefix else name
            access = "public" if safe(lambda p=prop: p.GetMethod.IsPublic, False) else "non-public"
            prop_type = safe(lambda p=prop: p.PropertyType.FullName, "")
            try:
                value = prop.GetValue(current, None)
                add_row(path, "Property", prop_type, get_type_name(value), value if is_scalar(value) else "", access, "")
                if (
                    value is not None
                    and not is_scalar(value)
                    and depth < max_depth
                    and should_recurse_member(name, path, value, depth)
                ):
                    walk(value, path, depth + 1)
            except Exception as exc:
                add_row(path, "Property", prop_type, "", "", access, exc)

        for field in safe(lambda: net_type.GetFields(BIND_ALL), []) or []:
            name = safe(lambda f=field: f.Name, "")
            if not name:
                continue
            path = f"{prefix}.{name}" if prefix else name
            access = "public" if safe(lambda f=field: f.IsPublic, False) else "non-public"
            field_type = safe(lambda f=field: f.FieldType.FullName, "")
            try:
                value = field.GetValue(current)
                add_row(path, "Field", field_type, get_type_name(value), value if is_scalar(value) else "", access, "")
                if (
                    value is not None
                    and not is_scalar(value)
                    and depth < max_depth
                    and should_recurse_member(name, path, value, depth)
                ):
                    walk(value, path, depth + 1)
            except Exception as exc:
                add_row(path, "Field", field_type, "", "", access, exc)

    walk(obj, root_name, 0)

    # Methods are listed at root level; recursive method dumps become too noisy.
    net_type = safe(lambda: obj.GetType(), None)
    if net_type is not None:
        for method in safe(lambda: net_type.GetMethods(BIND_ALL), []) or []:
            if safe(lambda m=method: m.IsSpecialName, False):
                continue
            access = "public" if safe(lambda m=method: m.IsPublic, False) else "non-public"
            add_row(
                safe(lambda m=method: m.Name, ""),
                "Method",
                method_signature(method),
                "",
                "",
                access,
                "",
            )
    return rows


def rows_to_value_map(rows: list[dict]) -> OrderedDict:
    values = OrderedDict()
    for row in rows:
        if row.get("MemberKind") not in {"Property", "Field", "Value"}:
            continue
        value = row.get("Value", "")
        if str(value).strip() == "":
            continue
        values[row.get("Path", "")] = value
    return values


def get_attr_any(obj: Any, names: Iterable[str], call_zero_arg: bool = False):
    for name in names:
        if obj is None:
            return None
        value = safe(lambda n=name: getattr(obj, n), None)
        if value is None:
            continue
        if call_zero_arg and callable(value):
            called = safe(lambda v=value: v(), None)
            if called is not None:
                return called
        elif not callable(value):
            return value
    return None


def call_method_any(obj: Any, names: Iterable[str]):
    for name in names:
        method = safe(lambda n=name: getattr(obj, n), None)
        if callable(method):
            value = safe(lambda m=method: m(), None)
            if value is not None:
                return value
    return None


def first_value(obj: Any, methods=(), attrs=(), phase_props=()):
    value = call_method_any(obj, methods)
    if value is not None:
        return value
    value = get_attr_any(obj, attrs)
    if value is not None:
        return value
    for phase, props, label in get_phase_items(obj):
        if label.lower() not in {"overall", "mixture", "mixed", "bulk", "0"}:
            continue
        value = get_attr_any(props, phase_props)
        if value is not None:
            return value
    return None


def get_phase_items(ms: Any):
    phases = safe(lambda: ms.Phases, None)
    result = []
    seen = set()

    # DWSIM material streams commonly expose the overall phase directly as
    # stream.Mixture. The PDF/report values for compound fractions are usually
    # under Mixture.Compounds[compound].MassFraction/MoleFraction.
    for attr, default_label in [
        ("Mixture", "Overall"),
        ("Overall", "Overall"),
        ("Vapor", "Vapor"),
        ("Liquid", "Liquid"),
        ("Liquid1", "Liquid1"),
        ("Liquid2", "Liquid2"),
        ("Solid", "Solid"),
    ]:
        phase = safe(lambda a=attr: getattr(ms, a), None)
        if phase is not None:
            oid = object_hash(phase)
            if oid not in seen:
                seen.add(oid)
                label = safe(lambda p=phase: p.Name, default_label) or default_label
                result.append((phase, safe(lambda p=phase: p.Properties, None), sanitize_name(label, default_label)))

    if phases is None:
        return result

    for idx, default_label in enumerate(["Overall", "Vapor", "Liquid1", "Liquid2", "Solid"]):
        phase = safe(lambda i=idx: phases[i], None)
        if phase is not None:
            oid = object_hash(phase)
            if oid not in seen:
                seen.add(oid)
                label = safe(lambda p=phase: p.Name, default_label) or default_label
                result.append((phase, safe(lambda p=phase: p.Properties, None), sanitize_name(label, default_label)))

    keys = safe(lambda: phases.Keys, None)
    if keys is not None:
        for key in keys:
            phase = safe(lambda k=key: phases[k], None)
            if phase is not None:
                oid = object_hash(phase)
                if oid not in seen:
                    seen.add(oid)
                    label = safe(lambda p=phase: p.Name, key) or key
                    result.append((phase, safe(lambda p=phase: p.Properties, None), sanitize_name(label, clean_text(key))))

    if not result:
        for key, phase in iter_net_items(phases, 20):
            if phase is not None:
                label = safe(lambda p=phase: p.Name, key) or key
                result.append((phase, safe(lambda p=phase: p.Properties, None), sanitize_name(label, key)))
    return result


def add_phase_properties(row: dict, ms: Any):
    prop_aliases = OrderedDict([
        ("Temperature_K", ["temperature", "Temperature"]),
        ("Pressure_Pa", ["pressure", "Pressure"]),
        ("MassFlow_kg_s", ["massflow", "massFlow", "MassFlow"]),
        ("MolarFlow_mol_s", ["molarflow", "molarFlow", "MolarFlow"]),
        ("VolumetricFlow_m3_s", ["volumetric_flow", "volumetricFlow", "VolumetricFlow"]),
        ("Density_kg_m3", ["density", "Density"]),
        ("Viscosity_Pa_s", ["viscosity", "Viscosity"]),
        ("Cp_J_kg_K", ["heatCapacityCp", "HeatCapacityCp", "Cp"]),
        ("Cv_J_kg_K", ["heatCapacityCv", "HeatCapacityCv", "Cv"]),
        ("ThermalConductivity_W_m_K", ["thermalConductivity", "ThermalConductivity"]),
        ("MolecularWeight", ["molecularWeight", "MolecularWeight"]),
        ("VaporFraction", ["vaporFraction", "VaporFraction", "mole_fraction"]),
        ("Enthalpy_J_kg", ["enthalpy", "Enthalpy"]),
        ("Entropy_J_kg_K", ["entropy", "Entropy"]),
        ("CompressibilityFactor", ["compressibilityFactor", "CompressibilityFactor"]),
        ("SurfaceTension_N_m", ["surfaceTension", "SurfaceTension"]),
    ])
    for phase, props, phase_label in get_phase_items(ms):
        if props is None:
            continue
        for out_name, aliases in prop_aliases.items():
            value = get_attr_any(props, aliases)
            row[f"{phase_label}_{out_name}"] = to_excel(value)

        # Add public phase properties that were not in the aliases.
        net_type = safe(lambda: props.GetType(), None)
        if net_type is not None:
            for prop in safe(lambda: net_type.GetProperties(BIND_PUBLIC), []) or []:
                if safe(lambda p=prop: p.GetIndexParameters().Length, 1) != 0:
                    continue
                name = safe(lambda p=prop: p.Name, "")
                key = f"{phase_label}_{sanitize_name(name)}"
                if key in row:
                    continue
                value = safe(lambda p=prop: p.GetValue(props, None), None)
                if is_scalar(value):
                    row[key] = to_excel(value)


def extract_material_stream(ms: Any, conns_by_tag: dict) -> dict:
    tag = get_tag(ms)
    temp = first_value(
        ms,
        methods=["GetTemperature"],
        attrs=["Temperature"],
        phase_props=["temperature", "Temperature"],
    )
    pressure = first_value(
        ms,
        methods=["GetPressure"],
        attrs=["Pressure"],
        phase_props=["pressure", "Pressure"],
    )
    mass_flow = first_value(
        ms,
        methods=["GetMassFlow"],
        attrs=["MassFlow"],
        phase_props=["massflow", "MassFlow"],
    )
    molar_flow = first_value(
        ms,
        methods=["GetMolarFlow"],
        attrs=["MolarFlow"],
        phase_props=["molarflow", "MolarFlow"],
    )
    vol_flow = first_value(
        ms,
        methods=["GetVolumetricFlow"],
        attrs=["VolumetricFlow"],
        phase_props=["volumetric_flow", "VolumetricFlow"],
    )

    row = {
        "Tag": tag,
        "ObjectType": "Material Stream",
        "Description": get_description(ms),
        "CalculationStatus": get_calc_status(ms),
        "ErrorMessage": get_error_message(ms),
        "ClassName": type(ms).__name__,
        "Temperature_K": to_excel(temp),
        "Temperature_C": k_to_c(temp),
        "Pressure_Pa": to_excel(pressure),
        "Pressure_bar": pa_to_bar(pressure),
        "MassFlow_kg_s": to_excel(mass_flow),
        "MolarFlow_mol_s": to_excel(molar_flow),
        "VolumetricFlow_m3_s": to_excel(vol_flow),
        "Density_kg_m3": to_excel(first_value(ms, methods=["GetDensity"], attrs=["Density"], phase_props=["density", "Density"])),
        "Viscosity_Pa_s": to_excel(first_value(ms, methods=["GetViscosity"], attrs=["Viscosity"], phase_props=["viscosity", "Viscosity"])),
        "Cp_J_kg_K": to_excel(first_value(ms, methods=["GetMassHeatCapacityCp"], attrs=["Cp"], phase_props=["heatCapacityCp", "Cp"])),
        "ThermalConductivity_W_m_K": to_excel(first_value(ms, methods=["GetThermalConductivity"], attrs=["ThermalConductivity"], phase_props=["thermalConductivity"])),
        "MolecularWeight": to_excel(first_value(ms, methods=["GetMolecularWeight"], attrs=["MolecularWeight"], phase_props=["molecularWeight"])),
        "VaporFraction": to_excel(first_value(ms, methods=["GetVaporFraction"], attrs=["VaporFraction"], phase_props=["vaporFraction", "mole_fraction"])),
        "Enthalpy_J_kg": to_excel(first_value(ms, methods=["GetEnthalpy"], attrs=["Enthalpy"], phase_props=["enthalpy"])),
        "Entropy_J_kg_K": to_excel(first_value(ms, methods=["GetEntropy"], attrs=["Entropy"], phase_props=["entropy"])),
        "InletConnections": sequence_to_text([c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Inlet"]),
        "OutletConnections": sequence_to_text([c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Outlet"]),
    }
    add_phase_properties(row, ms)
    return row


def extract_compositions(ms: Any) -> list[dict]:
    rows = []
    tag = get_tag(ms)
    emitted = set()
    for phase, props, phase_label in get_phase_items(ms):
        compounds = safe(lambda p=phase: p.Compounds, None)
        if compounds is None:
            continue
        for compound_name, compound in iter_net_items(compounds, 1000):
            compound_label = (
                clean_text(get_attr_any(compound, ["ComponentName", "Name"]), 250)
                or clean_text(compound_name, 250)
            )
            key = (tag, phase_label, compound_label)
            if key in emitted:
                continue
            emitted.add(key)
            rows.append({
                "Tag": tag,
                "Phase": phase_label,
                "Compound": compound_label,
                "IsCOMPOUND": "YES" if compound_label.strip().lower() == "composition" else "",
                "MoleFraction": to_excel(get_attr_any(compound, ["MoleFraction", "molefraction", "MolarFraction"])),
                "MassFraction": to_excel(get_attr_any(compound, ["MassFraction", "massfraction"])),
                "MolarFlow_mol_s": to_excel(get_attr_any(compound, ["MolarFlow", "molarflow"])),
                "MassFlow_kg_s": to_excel(get_attr_any(compound, ["MassFlow", "massflow"])),
                "FugacityCoeff": to_excel(get_attr_any(compound, ["FugacityCoeff", "fugacityCoeff"])),
                "ActivityCoeff": to_excel(get_attr_any(compound, ["ActivityCoeff", "activityCoeff"])),
            })
    return rows


def extract_overall_compositions_from_value_map(tag: str, value_map: OrderedDict) -> list[dict]:
    """
    Fallback for DWSIM/pythonnet interface wrappers.

    Some IMaterialStream objects expose composition values to .NET reflection
    but not through normal Python getattr/indexing. The reliable reflected paths
    look like:
        Mixture.Compounds[Ethanol].MassFraction
        Mixture.Compounds[Ethanol].MoleFraction
    """
    compound_data = OrderedDict()
    # Match any phase prefix: Mixture, Overall, Phases[0], Phases[1], Phases[2], …
    pattern = re.compile(
        r"^(?:Mixture|Overall|Phases\[(?:\d+|Overall|Mixture|Vapor|Liquid\d*)\])"
        r"\.Compounds\[(?P<compound>[^\]]+)\]\.(?P<prop>_?[A-Za-z0-9]+)$",
        re.IGNORECASE,
    )

    prop_map = {
        "componentname": "Compound",
        "name": "Compound",
        "molefraction": "MoleFraction",
        "molarfraction": "MoleFraction",
        "massfraction": "MassFraction",
        "molarflow": "MolarFlow_mol_s",
        "massflow": "MassFlow_kg_s",
        "fugacitycoeff": "FugacityCoeff",
        "activitycoeff": "ActivityCoeff",
    }

    for path, value in value_map.items():
        match = pattern.match(path)
        if not match:
            continue
        raw_compound = clean_text(match.group("compound"), 250)
        raw_prop = match.group("prop")
        prop = raw_prop.lstrip("_").lower()
        out_key = prop_map.get(prop)
        if not out_key:
            continue

        row = compound_data.setdefault(raw_compound, {
            "Tag": tag,
            "Phase": "Overall",
            "Compound": raw_compound,
            "IsEthanol": "YES" if raw_compound.strip().lower() == "ethanol" else "",
            "MoleFraction": "",
            "MassFraction": "",
            "MolarFlow_mol_s": "",
            "MassFlow_kg_s": "",
            "FugacityCoeff": "",
            "ActivityCoeff": "",
        })

        # Prefer public property paths over private backing fields when both exist.
        if raw_prop.startswith("_") and str(row.get(out_key, "")).strip():
            continue
        row[out_key] = to_excel(value)
        if out_key == "Compound" and value:
            compound_name = clean_text(value, 250)
            row["Compound"] = compound_name
            row["IsEthanol"] = "YES" if compound_name.strip().lower() == "ethanol" else ""

    return list(compound_data.values())


def extract_energy_stream(es: Any, conns_by_tag: dict) -> dict:
    tag = get_tag(es)
    energy = get_attr_any(es, ["EnergyFlow", "Q", "Duty", "Power"])
    related = conns_by_tag.get(tag, [])
    return {
        "Tag": tag,
        "ObjectType": "Energy Stream",
        "Description": get_description(es),
        "CalculationStatus": get_calc_status(es),
        "ErrorMessage": get_error_message(es),
        "ClassName": type(es).__name__,
        "EnergyFlow_W": to_excel(energy * 1000.0 if energy is not None else None),
        "EnergyFlow_kW": to_excel(energy),
        "AttachedEquipment": sequence_to_text([c["OtherTag"] for c in related]),
        "Direction": sequence_to_text([c["Direction"] for c in related]),
    }


def list_attr_values(obj: Any, names: Iterable[str]) -> list:
    for name in names:
        value = get_attr_any(obj, [name])
        if value is None:
            continue
        if isinstance(value, (list, tuple)):
            return list(value)
        if is_iterable_net(value):
            return [item for _, item in iter_net_items(value, 2000) if is_scalar(item)]
        if is_scalar(value):
            return [value]
    return []


def get_from_value_map(value_map: OrderedDict | None, names: Iterable[str]):
    if not value_map:
        return None
    wanted = [name.lower() for name in names]
    candidates = []
    for path, value in value_map.items():
        p = path.lower()
        for name in wanted:
            if p == name or p.endswith(f".{name}"):
                candidates.append((path.count(".") + path.count("["), path, value))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], len(item[1])))
    return candidates[0][2]


def get_any_value(obj: Any, names: Iterable[str], value_map: OrderedDict | None = None):
    value = get_attr_any(obj, names)
    if value is not None:
        return value
    return get_from_value_map(value_map, names)


def extract_profile_from_map(value_map: OrderedDict, include: Iterable[str], exclude: Iterable[str] = ()) -> list:
    include_l = [x.lower() for x in include]
    exclude_l = [x.lower() for x in exclude]
    matches = []
    index_re = re.compile(r"(?:\[|\.)(\d+)(?:\]|$)")
    for path, value in value_map.items():
        p = path.lower()
        if not all(x in p for x in include_l):
            continue
        if any(x in p for x in exclude_l):
            continue
        if not is_number(value):
            continue
        m = index_re.search(path)
        idx = int(m.group(1)) if m else len(matches)
        matches.append((idx, value))
    matches.sort(key=lambda x: x[0])
    return [v for _, v in matches]


def extract_stage_values(value_map: OrderedDict, aliases: Iterable[str]) -> list:
    aliases_l = {alias.lower() for alias in aliases}
    pattern = re.compile(r"^Stages\[(\d+)\]\.(.+)$", re.IGNORECASE)
    values = {}
    for path, value in value_map.items():
        match = pattern.match(path)
        if not match:
            continue
        tail = match.group(2).lower()
        if tail not in aliases_l:
            continue
        if not is_number(value):
            continue
        stage_index = int(match.group(1))
        values[stage_index] = value
    return [values[i] for i in sorted(values)]


def extract_distillation_column(col: Any, conns_by_tag: dict, value_map: OrderedDict) -> dict:
    tag = get_tag(col)

    # Try parsing ColumnPropertiesProfile
    profile_str = getattr(col, "ColumnPropertiesProfile", "") or ""
    parsed_t, parsed_p, parsed_l, parsed_v = [], [], [], []
    if profile_str:
        for line in profile_str.split("\n"):
            line = line.strip()
            parts = line.split()
            if parts and parts[0].isdigit():
                try:
                    # parts: [stage, p_bar, t_c, mv_kmol_h, ..., ml_kmol_h, ...]
                    p_pa = float(parts[1]) * 1e5
                    t_k = float(parts[2]) + 273.15
                    mv_mol_s = float(parts[3]) * 1000.0 / 3600.0
                    ml_mol_s = float(parts[8]) * 1000.0 / 3600.0
                    parsed_t.append(t_k)
                    parsed_p.append(p_pa)
                    parsed_v.append(mv_mol_s)
                    parsed_l.append(ml_mol_s)
                except:
                    pass

    stage_t = parsed_t or list_attr_values(col, ["StageTemperatures", "StageTemperature", "Temperatures", "Tc", "_tc"])
    stage_p = parsed_p or list_attr_values(col, ["StagePressures", "StagePressure", "Pressures", "Pc", "_pc"])
    stage_l = parsed_l or list_attr_values(col, ["StageLiquidMolarFlows", "LiquidMolarFlows", "LiquidFlows", "Lc", "_lc"])
    stage_v = parsed_v or list_attr_values(col, ["StageVaporMolarFlows", "VaporMolarFlows", "VaporFlows", "Vc", "_vc"])
    stage_e = list_attr_values(col, ["StageEfficiencies", "MurphreeEfficiencies", "Murphree_Eff", "Efficiencies"])

    if not stage_t:
        stage_t = extract_stage_values(value_map, ["T", "Temperature"]) or extract_profile_from_map(value_map, ["tc"])
    if not stage_p:
        stage_p = extract_stage_values(value_map, ["P", "Pressure"]) or extract_profile_from_map(value_map, ["pc"])
    if not stage_l:
        stage_l = (
            extract_stage_values(value_map, ["L", "L.Value", "Lout.Value", "Lss.Value", "LiquidFlow", "LiquidMolarFlow"])
            or extract_profile_from_map(value_map, ["lc"])
        )
    if not stage_v:
        stage_v = (
            extract_stage_values(value_map, ["V", "V.Value", "Vout.Value", "Vss.Value", "VaporFlow", "VaporMolarFlow"])
            or extract_profile_from_map(value_map, ["vc"])
        )
    if not stage_e:
        stage_e = extract_stage_values(value_map, ["Efficiency", "Eff"]) or extract_profile_from_map(value_map, ["eff"])

    if not stage_e and stage_t:
        stage_e = [1.0] * len(stage_t)

    inlets = [c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Inlet"]
    outlets = [c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Outlet"]

    cond_duty_val = get_any_value(col, ["CondenserDuty", "Qc", "CondenserHeatDuty"], value_map)
    rebo_duty_val = get_any_value(col, ["ReboilerDuty", "Qr", "ReboilerHeatDuty"], value_map)

    row = {
        "Tag": tag,
        "ObjectType": classify(col),
        "Description": get_description(col),
        "CalculationStatus": get_calc_status(col),
        "ErrorMessage": get_error_message(col),
        "ClassName": type(col).__name__,
        "NumberOfStages": to_excel(get_any_value(col, ["NumberOfStages", "NStages"], value_map)),
        "FeedStage": to_excel(get_any_value(col, ["FeedStageNumber", "FeedStage", "FeedTray", "FeedStageIndex"], value_map)),
        "CondenserDuty_W": to_excel(cond_duty_val * 1000.0 if cond_duty_val is not None else None),
        "CondenserDuty_kW": to_excel(cond_duty_val),
        "ReboilerDuty_W": to_excel(rebo_duty_val * 1000.0 if rebo_duty_val is not None else None),
        "ReboilerDuty_kW": to_excel(rebo_duty_val),
        "CondenserPressure_Pa": to_excel(get_any_value(col, ["CondenserPressure", "TopPressure"], value_map)),
        "CondenserPressure_bar": pa_to_bar(get_any_value(col, ["CondenserPressure", "TopPressure"], value_map)),
        "ReboilerPressure_Pa": to_excel(get_any_value(col, ["ReboilerPressure", "BottomPressure"], value_map)),
        "ReboilerPressure_bar": pa_to_bar(get_any_value(col, ["ReboilerPressure", "BottomPressure"], value_map)),
        "RefluxRatio": to_excel(get_any_value(col, ["RefluxRatio", "RR"], value_map)),
        "BoilupRatio": to_excel(get_any_value(col, ["BoilupRatio", "BR"], value_map)),
        "ColumnPressureDrop_Pa": to_excel(get_any_value(col, ["ColumnPressureDrop", "DeltaP", "PressureDrop"], value_map)),
        "EstimatedDiameter_m": to_excel(get_any_value(col, ["EstimatedDiameter", "EstimatedTrayDiameter", "Diameter", "ColumnDiameter"], value_map)),
        "EstimatedHeight_m": to_excel(get_any_value(col, ["EstimatedHeight", "EstimatedTrayHeight", "Height", "ColumnHeight"], value_map)),
        "TraySpacing_m": to_excel(get_any_value(col, ["TraySpacing", "PlateSpacing"], value_map)),
        "CondenserSpecification": clean_text(get_any_value(col, ["CondenserSpecification", "CondenserSpec", "CondenserType"], value_map)),
        "ReboilerSpecification": clean_text(get_any_value(col, ["ReboilerSpecification", "ReboilerSpec", "ReboilerType"], value_map)),
        "StageTemperatures_K": sequence_to_text(stage_t),
        "StagePressures_Pa": sequence_to_text(stage_p),
        "StageEfficiencies": sequence_to_text(stage_e),
        "StageVaporFlows": sequence_to_text(stage_v),
        "StageLiquidFlows": sequence_to_text(stage_l),
        "ConnectedFeedStreams": sequence_to_text(inlets),
        "ConnectedProductStreams": sequence_to_text(outlets),
        "HiddenGeometryVariables": hidden_values_text(value_map, ["diameter", "height", "spacing", "tray", "sizing", "geometry"]),
        "SizingVariables": hidden_values_text(value_map, ["sizing", "area", "volume", "weight", "diameter", "height"]),
        "MechanicalDesignVariables": hidden_values_text(value_map, ["mechanical", "wall", "thickness", "weight", "design"]),
    }
    row["_stage_t"] = stage_t
    row["_stage_p"] = stage_p
    row["_stage_l"] = stage_l
    row["_stage_v"] = stage_v
    row["_stage_e"] = stage_e
    return row


def hidden_values_text(value_map: OrderedDict, keywords: Iterable[str], limit: int = 40) -> str:
    keys = [k.lower() for k in keywords]
    found = []
    for path, value in value_map.items():
        p = path.lower()
        if any(k in p for k in keys) and str(value).strip():
            found.append(f"{path}={clean_text(value, 120)}")
            if len(found) >= limit:
                break
    return " | ".join(found)


def column_stage_rows(column_row: dict) -> list[dict]:
    tag = column_row.get("Tag", "")
    temps = column_row.get("_stage_t", []) or []
    press = column_row.get("_stage_p", []) or []
    liq = column_row.get("_stage_l", []) or []
    vap = column_row.get("_stage_v", []) or []
    eff = column_row.get("_stage_e", []) or []
    n = max(len(temps), len(press), len(liq), len(vap), len(eff))
    rows = []
    for i in range(n):
        t = temps[i] if i < len(temps) else None
        p = press[i] if i < len(press) else None
        rows.append({
            "Tag": tag,
            "Stage": i + 1,
            "Temperature_K": to_excel(t),
            "Temperature_C": k_to_c(t),
            "Pressure_Pa": to_excel(p),
            "Pressure_bar": pa_to_bar(p),
            "LiquidFlow": to_excel(liq[i] if i < len(liq) else None),
            "VaporFlow": to_excel(vap[i] if i < len(vap) else None),
            "Efficiency": to_excel(eff[i] if i < len(eff) else None),
        })
    return rows


def extract_heat_exchanger(hx: Any, conns_by_tag: dict, value_map: OrderedDict) -> dict:
    tag = get_tag(hx)
    duty = get_any_value(hx, ["Duty", "HeatDuty", "EnergyImbalance", "Q", "DeltaQ"], value_map)
    return {
        "Tag": tag,
        "ObjectType": classify(hx),
        "Description": get_description(hx),
        "CalculationStatus": get_calc_status(hx),
        "ErrorMessage": get_error_message(hx),
        "ClassName": type(hx).__name__,
        "Duty_W": to_excel(duty * 1000.0 if duty is not None else None),
        "Duty_kW": to_excel(duty),
        "Area_m2": to_excel(get_any_value(hx, ["Area", "HeatTransferArea", "HeatExchangerArea"], value_map)),
        "U_W_m2_K": to_excel(get_any_value(hx, ["OverallHeatTransferCoefficient", "U", "UValue"], value_map)),
        "LMTD_K": to_excel(get_any_value(hx, ["LMTD", "LogMeanTemperatureDifference"], value_map)),
        "NTU": to_excel(get_any_value(hx, ["NTU"], value_map)),
        "HotSideConditions": hidden_values_text(value_map, ["hot"]),
        "ColdSideConditions": hidden_values_text(value_map, ["cold"]),
        "HotSidePressureDrop_Pa": to_excel(get_any_value(hx, ["HotSidePressureDrop", "HotSideDeltaP"], value_map)),
        "ColdSidePressureDrop_Pa": to_excel(get_any_value(hx, ["ColdSidePressureDrop", "ColdSideDeltaP"], value_map)),
        "PressureDrops": hidden_values_text(value_map, ["pressuredrop", "delta_p", "deltap"]),
        "TubeLength_m": to_excel(get_any_value(hx, ["TubeLength"], value_map)),
        "TubeDiameter_m": to_excel(get_any_value(hx, ["TubeDiameter"], value_map)),
        "TubeCount": to_excel(get_any_value(hx, ["TubeCount", "NumberOfTubes"], value_map)),
        "ShellDiameter_m": to_excel(get_any_value(hx, ["ShellDiameter"], value_map)),
    }


def extract_pump(pump: Any, conns_by_tag: dict, value_map: OrderedDict) -> dict:
    power = get_any_value(pump, ["ActualPower", "Power", "EnergyFlow", "Work"], value_map)
    dp = get_any_value(pump, ["DeltaP", "PressureIncrease", "PressureRise"], value_map)
    return {
        "Tag": get_tag(pump),
        "ObjectType": classify(pump),
        "Description": get_description(pump),
        "CalculationStatus": get_calc_status(pump),
        "ErrorMessage": get_error_message(pump),
        "ClassName": type(pump).__name__,
        "Power_W": to_excel(power * 1000.0 if power is not None else None),
        "Power_kW": to_excel(power),
        "Efficiency": to_excel(get_any_value(pump, ["AdiabaticEfficiency", "Efficiency", "EfficiencyPump"], value_map)),
        "PressureRise_Pa": to_excel(dp),
        "PressureRise_bar": pa_to_bar(dp),
        "OutletPressure_Pa": to_excel(get_any_value(pump, ["OutletPressure"], value_map)),
        "NPSH": to_excel(get_any_value(pump, ["NPSH", "NPSHavailable", "NPSHrequired"], value_map)),
        "FlowRate": to_excel(get_any_value(pump, ["FlowRate", "VolumetricFlow"], value_map)),
        "HeadDeveloped_m": to_excel(get_any_value(pump, ["HeadDeveloped", "Head"], value_map)),
        "SizingVariables": hidden_values_text(value_map, ["npsh", "head", "flow", "power", "efficiency"]),
    }


def extract_compressor(comp: Any, conns_by_tag: dict, value_map: OrderedDict) -> dict:
    power = get_any_value(comp, ["ActualPower", "Power", "EnergyFlow", "Work"], value_map)
    return {
        "Tag": get_tag(comp),
        "ObjectType": classify(comp),
        "Description": get_description(comp),
        "CalculationStatus": get_calc_status(comp),
        "ErrorMessage": get_error_message(comp),
        "ClassName": type(comp).__name__,
        "Power_W": to_excel(power * 1000.0 if power is not None else None),
        "Power_kW": to_excel(power),
        "Efficiency": to_excel(get_any_value(comp, ["AdiabaticEfficiency", "PolytropicEfficiency", "Efficiency"], value_map)),
        "PressureRatio": to_excel(get_any_value(comp, ["PressureRatio", "CompressionRatio"], value_map)),
        "DeltaP_Pa": to_excel(get_any_value(comp, ["DeltaP"], value_map)),
        "OutletPressure_Pa": to_excel(get_any_value(comp, ["OutletPressure"], value_map)),
        "AdiabaticHead": to_excel(get_any_value(comp, ["AdiabaticHead"], value_map)),
        "PolytropicHead": to_excel(get_any_value(comp, ["PolytropicHead"], value_map)),
        "SizingVariables": hidden_values_text(value_map, ["head", "ratio", "power", "efficiency", "flow"]),
    }


def extract_generic_equipment(obj: Any, conns_by_tag: dict, value_map: OrderedDict) -> dict:
    tag = get_tag(obj)
    row = {
        "Tag": tag,
        "ObjectType": classify(obj),
        "Description": get_description(obj),
        "CalculationStatus": get_calc_status(obj),
        "ErrorMessage": get_error_message(obj),
        "ClassName": type(obj).__name__,
        "ConnectedInlets": sequence_to_text([c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Inlet"]),
        "ConnectedOutlets": sequence_to_text([c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Outlet"]),
    }
    for attr in [
        "Volume", "Diameter", "Height", "Length", "Area", "Duty", "Power",
        "EnergyImbalance", "HeatAdded", "DeltaP", "DeltaT", "OutletPressure",
        "OutletTemperature", "Efficiency", "ResidenceTime", "Conversion",
    ]:
        value = get_any_value(obj, [attr], value_map)
        if value is not None:
            row[attr] = to_excel(value)
    row["CostingLikeVariables"] = hidden_values_text(value_map, COSTING_KEYWORDS)
    return row


def connector_tag(connector: Any, attrs: Iterable[str]) -> str:
    for attr in attrs:
        value = safe(lambda a=attr: getattr(connector, a), None)
        if value is None:
            continue
        tag = safe(lambda v=value: v.Tag, None) or safe(lambda v=value: v.Name, None)
        if tag:
            return clean_text(tag, 200)
        if is_scalar(value):
            return clean_text(value, 200)
    return ""


def build_connection_graph(sim: Any) -> list[dict]:
    rows = []
    objects = safe(lambda: sim.SimulationObjects, None)
    if objects is None:
        return rows

    for key, obj in iter_net_items(objects, 10000):
        tag = get_tag(obj, key)
        go = safe(lambda o=obj: o.GraphicObject, None)
        if go is None:
            continue
        for side_name, direction, connectors_attr in [
            ("Input", "Inlet", "InputConnectors"),
            ("Output", "Outlet", "OutputConnectors"),
            ("Energy", "Energy", "EnergyConnector"),
            ("EnergyInput", "Energy Inlet", "EnergyInputConnectors"),
            ("EnergyOutput", "Energy Outlet", "EnergyOutputConnectors"),
        ]:
            connectors = safe(lambda a=connectors_attr: getattr(go, a), None)
            if connectors is None:
                continue
            if not is_iterable_net(connectors):
                connectors = [connectors]
            for idx, conn in iter_net_items(connectors, 100):
                attached = safe(lambda c=conn: c.IsAttached, None)
                attached_conn = safe(lambda c=conn: c.AttachedConnector, None)
                if attached is False and attached_conn is None:
                    continue

                stream_tag = (
                    connector_tag(conn, ["AttachedConnector", "AttachedTo", "AttachedFrom"])
                    or clean_text(safe(lambda c=conn: c.Name, ""), 200)
                )
                other_tag = connector_tag(attached_conn, ["AttachedFrom", "AttachedTo"]) if attached_conn is not None else ""
                if not other_tag:
                    other_tag = connector_tag(conn, ["AttachedFrom", "AttachedTo"])
                if not stream_tag and not other_tag:
                    continue

                if direction.startswith("Inlet"):
                    from_tag, to_tag = other_tag or stream_tag, tag
                elif direction.startswith("Outlet"):
                    from_tag, to_tag = tag, other_tag or stream_tag
                else:
                    from_tag, to_tag = tag, other_tag or stream_tag

                rows.append({
                    "From": from_tag,
                    "To": to_tag,
                    "StreamTag": stream_tag,
                    "ObjectTag": tag,
                    "OtherTag": other_tag,
                    "Direction": direction,
                    "ConnectorSide": side_name,
                    "ConnectorIndex": idx,
                    "ConnectorName": clean_text(safe(lambda c=conn: c.Name, ""), 200),
                })
    return rows


def index_connections(rows: list[dict]) -> dict:
    by_tag = defaultdict(list)
    for row in rows:
        for key in ("ObjectTag", "StreamTag", "From", "To"):
            tag = row.get(key, "")
            if tag:
                by_tag[tag].append(row)
    return by_tag


def is_costing_key(key: str) -> bool:
    text = key.lower()
    return any(keyword in text for keyword in COSTING_KEYWORDS)


def cost_relevance(key: str) -> str:
    text = key.lower()
    return "CAPEX" if any(keyword in text for keyword in CAPEX_HINTS) else "OPEX"


def add_costing_rows(rows: list[dict], seen: set, tag: str, obj_type: str, source: str, values: dict):
    for key, value in values.items():
        if not is_costing_key(str(key)):
            continue
        text = clean_text(value, 1000).strip()
        if text in {"", "None", "[]", "{}", "0", "0.0"}:
            continue
        sig = (tag, obj_type, str(key), text, source)
        if sig in seen:
            continue
        seen.add(sig)
        rows.append({
            "Tag": tag,
            "ObjectType": obj_type,
            "Variable": str(key),
            "Value": text,
            "Source": source,
            "CostRelevance": cost_relevance(str(key)),
        })


def discover_dwsim_path(cli_path: str | None) -> str:
    candidates = []
    if cli_path:
        candidates.append(cli_path)
    for env_name in ("DWSIM_PATH", "DWSIM_HOME"):
        if os.environ.get(env_name):
            candidates.append(os.environ[env_name])
    for env_name in ("LOCALAPPDATA", "ProgramFiles", "ProgramFiles(x86)"):
        root = os.environ.get(env_name)
        if root:
            candidates.append(os.path.join(root, "DWSIM"))
            candidates.append(os.path.join(root, "DWSIM8"))
            candidates.append(os.path.join(root, "DWSIM7"))

    for path in candidates:
        if path and os.path.exists(os.path.join(path, "DWSIM.Automation.dll")):
            return os.path.abspath(path)
    raise FileNotFoundError(
        "Could not find DWSIM.Automation.dll. Pass --dwsim or set DWSIM_PATH."
    )


def load_dwsim(dwsim_path: str):
    global System, BindingFlags, RuntimeHelpers, BIND_PUBLIC, BIND_ALL

    # FIX #29: Insert DWSIM path at front of sys.path BEFORE importing clr.
    # ThermoCS and other property-package DLLs are resolved relative to sys.path;
    # if the DWSIM directory appears too late, CLR cannot find them.
    if dwsim_path not in sys.path:
        sys.path.insert(0, dwsim_path)
    # Also walk subdirs to register DLL directories, but do NOT add them to sys.path
    # to avoid polluting Python standard library paths with DWSIM's internal/bundled Python libraries.
    for root, dirs, files in os.walk(dwsim_path):
        if hasattr(os, "add_dll_directory"):
            os.add_dll_directory(root)

    os.environ["PATH"] = dwsim_path + os.pathsep + os.environ.get("PATH", "")

    import clr  # noqa: E402  (must be AFTER sys.path is set)

    preferred = [
        "ThermoCS",                  # FIX #29: load early so it is available to Thermodynamics
        "DWSIM.Interfaces",
        "DWSIM.SharedClasses",
        "DWSIM.Thermodynamics",
        "DWSIM.UnitOperations",
        "DWSIM.Automation",
    ]
    load_errors = []
    for name in preferred:
        # Try both full path and by name (for GAC/in-path assemblies)
        dll = os.path.join(dwsim_path, f"{name}.dll")
        if os.path.exists(dll):
            try:
                clr.AddReference(dll)
                continue
            except Exception as exc:
                load_errors.append(f"{name} (by path): {exc}")
        # Fallback: try by name alone (CLR resolves from sys.path)
        try:
            clr.AddReference(name)
        except Exception as exc:
            load_errors.append(f"{name} (by name): {exc}")

    if load_errors:
        print("Assembly load warnings:")
        for err in load_errors:
            print(f"  {err}")


    import System as _System
    from System.Reflection import BindingFlags as _BindingFlags
    from System.Runtime.CompilerServices import RuntimeHelpers as _RuntimeHelpers

    System = _System
    BindingFlags = _BindingFlags
    RuntimeHelpers = _RuntimeHelpers
    BIND_PUBLIC = BindingFlags.Public | BindingFlags.Instance | BindingFlags.Static
    BIND_ALL = (
        BindingFlags.Public
        | BindingFlags.NonPublic
        | BindingFlags.Instance
        | BindingFlags.Static
    )

    from System import AppDomain, ResolveEventHandler
    from System.Reflection import Assembly

    def assembly_resolve_handler(sender, args):
        dll_name = args.Name.split(',')[0] + ".dll"
        for root, dirs, files in os.walk(dwsim_path):
            if dll_name in files:
                try:
                    return Assembly.LoadFrom(os.path.join(root, dll_name))
                except Exception:
                    pass
        return None

    AppDomain.CurrentDomain.AssemblyResolve += ResolveEventHandler(assembly_resolve_handler)

    from DWSIM.Automation import Automation3
    return Automation3


def get_simulation_objects(sim: Any):
    objects = safe(lambda: sim.SimulationObjects.Values, None)
    if objects is None:
        return []
    
    res = []
    for key, obj in iter_net_items(objects, 100000):
        if hasattr(obj, "GetAsObject"):
            try:
                obj = obj.GetAsObject()
            except:
                pass
        res.append((key, obj))
    return res


def workbook_headers(rows: list[dict]) -> list[str]:
    headers = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen and not str(key).startswith("_"):
                seen.add(key)
                headers.append(key)
    return headers or ["No data"]


def write_sheet(wb, name: str, rows: list[dict], color: str = "1F4E79"):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    safe_name = re.sub(r"[:\\/?*\[\]]", "_", name)[:31]
    ws = wb.create_sheet(safe_name)
    headers = workbook_headers(rows)
    ws.append(headers)

    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        ws.append(["No data"])
    else:
        alt = PatternFill("solid", fgColor="F4F8FB")
        for idx, row in enumerate(rows, start=2):
            ws.append([to_excel(row.get(h, "")) for h in headers])
            if idx % 2 == 0:
                for cell in ws[idx]:
                    cell.fill = alt
                    cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        width = min(max(12, len(str(header)) + 2), 60)
        sample = rows[:200] if rows else []
        for row in sample:
            width = min(max(width, len(clean_text(row.get(header, ""), 80)) + 2), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def write_summary(wb, summary: dict):
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Summary", 0)
    ws.append(["DWSIM Costing Introspection Report", ""])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    for key, value in summary.items():
        ws.append([key, value])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    return ws


def parse_args():
    parser = argparse.ArgumentParser(description="Extract DWSIM flowsheet data for CAPEX/OPEX costing.")
    parser.add_argument("flowsheet", help="Path to a .dwxmz flowsheet.")
    parser.add_argument("--output", "-o", help="Output Excel workbook path.")
    parser.add_argument("--dwsim", help="DWSIM installation folder containing DWSIM.Automation.dll.")
    parser.add_argument("--depth", type=int, default=MAX_DISCOVERY_DEPTH, help="Recursive discovery depth.")
    parser.add_argument("--collection-limit", type=int, default=MAX_COLLECTION_ITEMS, help="Max items per collection.")
    parser.add_argument("--no-calc", action="store_true", help="Skip flowsheet recalculation (useful if flowsheet is already solved).")
    return parser.parse_args()


def main():
    global MAX_COLLECTION_ITEMS

    args = parse_args()
    flowsheet = os.path.abspath(args.flowsheet)
    if not os.path.exists(flowsheet):
        raise FileNotFoundError(flowsheet)
    if not flowsheet.lower().endswith(".dwxmz"):
        print("Warning: input does not end with .dwxmz; DWSIM may still load it if supported.")

    output = args.output or os.path.splitext(flowsheet)[0] + "_costing_introspection.xlsx"
    output = os.path.abspath(output)
    MAX_COLLECTION_ITEMS = args.collection_limit

    dwsim_path = discover_dwsim_path(args.dwsim)
    Automation3 = load_dwsim(dwsim_path)

    print("=" * 72)
    print("DWSIM costing introspection engine")
    print(f"DWSIM     : {dwsim_path}")
    print(f"Flowsheet : {flowsheet}")
    print(f"Output    : {output}")
    print("=" * 72)

    automation = Automation3()
    sim = automation.LoadFlowsheet(flowsheet)
    if not args.no_calc:
        print("Calculating flowsheet...")
        automation.CalculateFlowsheet2(sim)
        print("Flowsheet calculation complete.")
    else:
        print("Skipping flowsheet calculation (using pre-calculated values).")
    objects = get_simulation_objects(sim)
    print(f"Loaded flowsheet. Simulation objects discovered: {len(objects)}")

    connections = build_connection_graph(sim)
    conns_by_tag = index_connections(connections)

    objects_rows = []
    property_rows = []
    field_rows = []
    method_rows = []
    discovery = []
    streams = []
    compositions = []
    energy_streams = []
    columns = []
    column_stages = []
    heat_exchangers = []
    pumps = []
    compressors = []
    reactors = []
    other_equipment = []
    costing = []
    costing_seen = set()

    for key, obj in objects:
        tag = get_tag(obj, key)
        obj_type = classify(obj)
        class_name = type(obj).__name__
        full_type = get_type_name(obj)
        print(f"[{obj_type:<24}] {tag}")

        props, fields, methods = member_rows(obj, tag, obj_type, class_name)
        property_rows.extend(props)
        field_rows.extend(fields)
        method_rows.extend(methods)

        disc = discovery_rows(obj, tag, obj_type, "", max_depth=args.depth)
        discovery.extend(disc)
        value_map = rows_to_value_map(disc)

        inlets = [c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Inlet"]
        outlets = [c["StreamTag"] or c["OtherTag"] for c in conns_by_tag.get(tag, []) if c["Direction"] == "Outlet"]
        objects_rows.append({
            "Tag": tag,
            "Key": clean_text(key, 250),
            "ObjectType": obj_type,
            "ClassName": class_name,
            "FullTypeName": full_type,
            "Description": get_description(obj),
            "CalculationStatus": get_calc_status(obj),
            "ErrorMessage": get_error_message(obj),
            "ConnectedInletStreams": sequence_to_text(inlets),
            "ConnectedOutletStreams": sequence_to_text(outlets),
            "PublicPropertyCount": len(props),
            "PublicFieldCount": len(fields),
            "PublicMethodCount": len(methods),
        })

        class_text = f"{class_name} {full_type}"
        targeted_row = None
        if "MaterialStream" in class_text:
          targeted_row = extract_material_stream(obj, conns_by_tag)
          streams.append(targeted_row)

          comp_rows = extract_compositions(obj)

          if not comp_rows:
              comp_rows = extract_overall_compositions_from_value_map(
                  tag,
                  value_map,
              )

          if not comp_rows and args.depth < 4:
              # Deep scan: temporarily go to depth=4 just for this stream
              comp_discovery = discovery_rows(
                  obj,
                  tag,
                  obj_type,
                  "",
                  max_depth=4,
              )
              comp_rows = extract_overall_compositions_from_value_map(
                  tag,
                  rows_to_value_map(comp_discovery),
              )

          compositions.extend(comp_rows)
          
        elif "EnergyStream" in class_text:
            targeted_row = extract_energy_stream(obj, conns_by_tag)
            energy_streams.append(targeted_row)
        elif any(x in class_text for x in ("DistillationColumn", "AbsorptionColumn", "ShortcutColumn")):
            targeted_row = extract_distillation_column(obj, conns_by_tag, value_map)
            columns.append({k: v for k, v in targeted_row.items() if not k.startswith("_")})
            column_stages.extend(column_stage_rows(targeted_row))
        elif any(x in class_text for x in ("HeatExchanger", "Heater", "Cooler")):
            targeted_row = extract_heat_exchanger(obj, conns_by_tag, value_map)
            heat_exchangers.append(targeted_row)
        elif "Pump" in class_text:
            targeted_row = extract_pump(obj, conns_by_tag, value_map)
            pumps.append(targeted_row)
        elif any(x in class_text for x in ("Compressor", "Expander")):
            targeted_row = extract_compressor(obj, conns_by_tag, value_map)
            compressors.append(targeted_row)
        else:
            targeted_row = extract_generic_equipment(obj, conns_by_tag, value_map)
            if "Reactor" in class_text:
                reactors.append(targeted_row)
            else:
                other_equipment.append(targeted_row)

        add_costing_rows(costing, costing_seen, tag, obj_type, "TargetedExtractor", targeted_row or {})
        add_costing_rows(costing, costing_seen, tag, obj_type, "RecursiveDiscovery", value_map)

    summary = OrderedDict([
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Flowsheet", flowsheet),
        ("DWSIMPath", dwsim_path),
        ("Objects", len(objects_rows)),
        ("Properties", len(property_rows)),
        ("Fields", len(field_rows)),
        ("Methods", len(method_rows)),
        ("MaterialStreams", len(streams)),
        ("Compositions", len(compositions)),
        ("EnergyStreams", len(energy_streams)),
        ("DistillationColumns", len(columns)),
        ("ColumnStageRows", len(column_stages)),
        ("HeatExchangers", len(heat_exchangers)),
        ("Pumps", len(pumps)),
        ("Compressors", len(compressors)),
        ("Reactors", len(reactors)),
        ("OtherEquipment", len(other_equipment)),
        ("Connections", len(connections)),
        ("CostingVariables", len(costing)),
        ("DiscoveryDumpRows", len(discovery)),
    ])

    print("Writing Excel workbook...")
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary(wb, summary)

    colors = {
        "Objects": "1F4E79",
        "Properties": "263238",
        "Fields": "37474F",
        "Methods": "455A64",
        "Streams": "2E7D32",
        "Compositions": "1B5E20",
        "EnergyStreams": "E65100",
        "DistillationColumns": "4A148C",
        "ColumnStageProfiles": "6A1B9A",
        "HeatExchangers": "B71C1C",
        "Pumps": "0D47A1",
        "Compressors": "01579B",
        "Reactors": "880E4F",
        "OtherEquipment": "33691E",
        "Connections": "37474F",
        "CostingVariables": "BF360C",
        "DiscoveryDump": "111111",
    }

    sheet_payloads = OrderedDict([
        ("Objects", objects_rows),
        ("Properties", property_rows),
        ("Fields", field_rows),
        ("Methods", method_rows),
        ("Streams", streams),
        ("Compositions", compositions),
        ("EnergyStreams", energy_streams),
        ("DistillationColumns", columns),
        ("ColumnStageProfiles", column_stages),
        ("HeatExchangers", heat_exchangers),
        ("Pumps", pumps),
        ("Compressors", compressors),
        ("Reactors", reactors),
        ("OtherEquipment", other_equipment),
        ("Connections", connections),
        ("CostingVariables", costing),
        ("Costing_Variables", costing),
        ("DiscoveryDump", discovery),
    ])

    for sheet in REQUIRED_SHEETS:
        if sheet not in sheet_payloads:
            sheet_payloads[sheet] = []

    for name, rows in sheet_payloads.items():
        write_sheet(wb, name, rows, colors.get(name, "1F4E79"))

    wb.save(output)
    print("=" * 72)
    print("Extraction complete")
    for key, value in summary.items():
        if key not in {"Flowsheet", "DWSIMPath"}:
            print(f"{key:<24}: {value}")
    print(f"Excel saved: {output}")
    print("=" * 72)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        sys.exit(1)
